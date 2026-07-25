# -*- coding: utf-8 -*-
"""Calibrage du filtre de similarité — PROUVE que la cascade est d'accord avec
le tri manuel AVANT de lui faire confiance. N'écrit RIEN d'automatique : affiche
la matrice d'accord, les désaccords, la distribution des scores et un balayage
de SIM_SEUIL_HAUT. C'est l'utilisatrice qui fixe ensuite les seuils.

    python calibrer_seuil.py

Nécessite pour le jugement RÉEL : GEMINI_API_KEY + le SDK google-generativeai.
Sans eux, chaque offre est marquée "A_VERIFIER (Gemini indisponible)" et le
tableau le signale clairement (aucun quota consommé, aucun plantage).
"""
import json
import os
import statistics
import sys
import unicodedata

import openpyxl

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import similarite as S

OUTDIR = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.join(OUTDIR, "Sourcing_regie_banque.xlsx")


def norm(s):
    s = unicodedata.normalize("NFD", str(s or "").lower())
    return "".join(c for c in s if unicodedata.category(c) != "Mn")


# --- NÉGATIFS connus (en dur, cf. spec §5) : (libellé, [descripteurs à matcher])
NEGATIFS_DESCR = [
    ("Bpifrance … DSI (CDI client final)", ["bpifrance", "dsi"]),
    ("VISEO … Consultant EPM … Rejoignez (CDI ESN)", ["viseo", "epm"]),
    ("Tectra … TMS Direction Financière (trésorerie corporate)", ["tectra", "tms"]),
]


def _charger_cache_textes():
    """url/titre -> annonce complète (pour récupérer le TEXTE intégral)."""
    idx = {}
    for pays in ("france", "maroc"):
        p = os.path.join(OUTDIR, f"cache_annonces_{pays}.json")
        try:
            with open(p, encoding="utf-8") as f:
                for a in json.load(f):
                    u = (a.get("url") or "").split("?")[0].rstrip("/")
                    if u:
                        idx[u] = a
                    idx.setdefault(norm(a.get("poste", "")), a)
        except FileNotFoundError:
            pass
    return idx


def charger_positifs(cache):
    """POSITIFS = lignes déjà classées ★★ / ★ dans mon fichier trié."""
    wb = openpyxl.load_workbook(XLSX)
    pos = []
    for tab in ["Maroc (sourcing)", "France (sourcing)"]:
        if tab not in wb.sheetnames:
            continue
        ws = wb[tab]
        last = max((r for r in range(2, ws.max_row + 1) if ws.cell(r, 2).value),
                   default=1)
        for r in range(2, last + 1):
            v = str(ws.cell(r, 12).value or "")
            if not (v.startswith("★★") or v.startswith("★")):
                continue
            poste = str(ws.cell(r, 2).value or "").lstrip("★ ").strip()
            entite = str(ws.cell(r, 4).value or "")
            mission = str(ws.cell(r, 3).value or "")
            url = None
            for c in range(1, ws.max_column + 1):
                if ws.cell(r, c).hyperlink:
                    url = ws.cell(r, c).hyperlink.target.split("?")[0].rstrip("/")
                    break
            a = cache.get(url) or cache.get(norm(poste)) or {}
            pos.append({
                "poste": poste, "mission": mission, "entite": entite,
                "texte": a.get("texte", ""),
                "label": "★★" if v.startswith("★★") else "★",
                "verdict_manuel": v,
            })
    return pos


def charger_negatifs(cache):
    """NÉGATIFS connus : cherchés dans le cache par descripteurs."""
    neg = []
    toutes = list(cache.values())
    for libelle, mots in NEGATIFS_DESCR:
        trouve = None
        for a in toutes:
            blob = norm(a.get("poste", "") + " " + a.get("entite", "") + " " +
                        a.get("texte", ""))
            if all(m in blob for m in mots):
                trouve = a
                break
        if trouve:
            neg.append({
                "poste": trouve.get("poste", ""), "mission": trouve.get("mission", ""),
                "entite": trouve.get("entite", ""), "texte": trouve.get("texte", ""),
                "label": "NÉG", "verdict_manuel": libelle,
            })
        else:
            print(f"  ! négatif introuvable dans le cache : {libelle}")
    return neg


def _med(xs):
    xs = [x for x in xs if x is not None]
    if not xs:
        return None
    return round(statistics.median(xs))


def afficher(rows):
    """rows = liste de (offre, resultat_similarite)."""
    pos = [(o, r) for o, r in rows if o["label"] in ("★★", "★")]
    neg = [(o, r) for o, r in rows if o["label"] == "NÉG"]
    indispo = sum(1 for _, r in rows if r["score_global"] is None)

    print("\n" + "═" * 78)
    print(f"  JEU DE CALIBRAGE : {len(pos)} positifs (★★/★)  ·  {len(neg)} négatifs connus")
    if indispo:
        print(f"  ⚠️  {indispo}/{len(rows)} offres NON jugées par Gemini "
              f"(clé/SDK absents) — matrice non fiable tant que ce n'est pas résolu.")
    print("═" * 78)

    # --- Distribution des scores par bande
    print("\n  DISTRIBUTION DES SCORES GEMINI (score_global 0-100)")
    for nom, sous in [("★★ MATCH CŒUR", [r for o, r in pos if o["label"] == "★★"]),
                      ("★  À SAISIR", [r for o, r in pos if o["label"] == "★"]),
                      ("NÉGATIFS", [r for _, r in neg])]:
        sc = [r["score_global"] for r in sous if r["score_global"] is not None]
        if sc:
            print(f"    {nom:16} n={len(sc):3}  min={min(sc):3}  "
                  f"médiane={_med(sc):3}  max={max(sc):3}")
        else:
            print(f"    {nom:16} (aucun score)")

    # --- Balayage de SIM_SEUIL_HAUT
    print("\n  BALAYAGE DE SIM_SEUIL_HAUT (rappel = mes positifs gardés ; "
          "fuites = négatifs gardés)")
    print(f"    {'seuil':>5} {'rappel ★★/★':>14} {'fuites nég.':>12}")
    pos_sc = [r["score_global"] for _, r in pos if r["score_global"] is not None]
    neg_sc = [r["score_global"] for _, r in neg if r["score_global"] is not None]
    for seuil in range(60, 86, 5):
        gardes_pos = sum(1 for s in pos_sc if s >= seuil)
        fuites = sum(1 for s in neg_sc if s >= seuil)
        rappel = f"{gardes_pos}/{len(pos_sc)}" if pos_sc else "-"
        pct = f"({100*gardes_pos//len(pos_sc)}%)" if pos_sc else ""
        print(f"    {seuil:>5} {rappel:>10} {pct:>4} {fuites:>8}/{len(neg_sc)}")

    # --- Désaccords à relire
    print("\n  DÉSACCORDS À RELIRE (Gemini diverge de mon tri)")
    seuil = S.SIM_SEUIL_HAUT if hasattr(S, "SIM_SEUIL_HAUT") else 75
    dis = 0
    for o, r in pos:
        s = r["score_global"]
        if s is not None and s < seuil:
            dis += 1
            print(f"    [MON {o['label']} → Gemini {s}] {o['poste'][:50]} "
                  f"({o['entite'][:18]})")
            print(f"        {r['verdict_gemini']} · {r.get('pole_principal','-')} · "
                  f"{r['raison'][:90]}")
    for o, r in neg:
        s = r["score_global"]
        if s is not None and s >= seuil:
            dis += 1
            print(f"    [NÉGATIF → Gemini {s}] {o['poste'][:50]} ({o['entite'][:18]})")
            print(f"        {r['verdict_gemini']} · {r['raison'][:90]}")
    if dis == 0:
        print("    (aucun désaccord au seuil courant)")

    print("\n  → Fixe ensuite SIM_SEUIL_HAUT / SIM_SEUIL_BAS d'après ce tableau.")
    print("    Objectif : garder ≥ 95 % de tes ★★ tout en coupant les négatifs.\n")


def main():
    cache = _charger_cache_textes()
    pos = charger_positifs(cache)
    neg = charger_negatifs(cache)
    offres = pos + neg
    if not offres:
        print("Aucune offre de calibrage trouvée (fichier trié vide ?).")
        return
    print(f"Jugement de {len(offres)} offres par la cascade "
          f"(embeddings si dispo + Gemini)…")
    resultats = S.similarite(offres, use_embeddings=True)
    afficher(list(zip(offres, resultats)))


if __name__ == "__main__":
    main()
