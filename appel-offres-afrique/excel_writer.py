"""Écriture/mise à jour incrémentale de l'Excel de veille Afrique.

Deux onglets, même logique que APPEL_OFFRES/excel_writer.py :
- "Appels d'offres" : avis actifs DGMP-CI + PMMP (source de référence).
- "Plans prévisionnels" : projets PPM (Côte d'Ivoire), non exhaustifs par
  nature (programme prévisionnel de l'acheteur, pas un avis ferme).

Dédoublonnage par identifiant (1ère colonne) — une ligne déjà présente est
mise à jour en place plutôt que dupliquée.
"""
from __future__ import annotations

import datetime as dt
import json
import logging
import time
from copy import copy
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet

import config

logger = logging.getLogger(__name__)

_SAVE_MAX_RETRIES = 5
_SAVE_RETRY_DELAY_SECONDS = 2.0

# ---------------------------------------------------------------------------
# Protection des modifications manuelles (demandé explicitement) — deux
# mécanismes complémentaires :
# 1) une ligne SURLIGNÉE (remplissage de cellule non vide, n'importe quelle
#    colonne) est considérée comme revue/décidée manuellement : ni ses
#    valeurs ni sa présence ne sont plus jamais modifiées automatiquement,
#    même si une collecte ultérieure la reclasserait "hors IT" ou ne la
#    retrouve plus dans la source.
# 2) une ligne SUPPRIMÉE manuellement (ID connu d'une exécution précédente,
#    absent de l'onglet actuel) n'est jamais réinjectée — un registre
#    persistant (`_ids_geres_<onglet>.json`) mémorise tous les ID déjà
#    proposés par le pipeline, pour distinguer "jamais vu" de "vu puis
#    supprimé volontairement".
# ---------------------------------------------------------------------------


def _ledger_path(ledger_dir: Path, sheet_name: str) -> Path:
    """Registre à côté du fichier Excel lui-même (même dossier que `path`),
    pas dans `config.DATA_DIR` en dur — sinon les tests (qui écrivent dans
    un `tmp_path` isolé) pollueraient le vrai dossier `data/` du projet."""
    safe = sheet_name.lower().replace(" ", "_").replace("'", "").replace("’", "")
    return ledger_dir / f"_ids_geres_{safe}.json"


def _load_ledger(ledger_dir: Path, sheet_name: str) -> set[str]:
    path = _ledger_path(ledger_dir, sheet_name)
    if not path.exists():
        return set()
    try:
        return set(json.loads(path.read_text(encoding="utf-8")))
    except (json.JSONDecodeError, OSError):
        return set()


def _save_ledger(ledger_dir: Path, sheet_name: str, ids: set[str]) -> None:
    _ledger_path(ledger_dir, sheet_name).write_text(
        json.dumps(sorted(ids), ensure_ascii=False, indent=2), encoding="utf-8"
    )


def _is_row_protected(ws: Worksheet, row_idx: int, num_columns: int) -> bool:
    """Une ligne est protégée dès qu'UNE de ses cellules porte un
    remplissage (surlignage) — signal manuel simple, sans nouvelle colonne
    ni convention à retenir : surligner une ligne dans Excel suffit."""
    for col_idx in range(1, num_columns + 1):
        fill = ws.cell(row=row_idx, column=col_idx).fill
        if fill is not None and fill.fill_type not in (None, "none"):
            return True
    return False


def _save_with_retry(wb: Workbook, path: Path) -> None:
    last_exc: OSError | None = None
    for attempt in range(_SAVE_MAX_RETRIES):
        try:
            wb.save(path)
            return
        except (PermissionError, OSError) as exc:
            last_exc = exc
            logger.warning(
                "Fichier Excel verrouillé (tentative %d/%d) : %s — nouvelle tentative dans %.0fs",
                attempt + 1, _SAVE_MAX_RETRIES, exc, _SAVE_RETRY_DELAY_SECONDS,
            )
            time.sleep(_SAVE_RETRY_DELAY_SECONDS)
    raise OSError(
        f"Impossible d'écrire {path} après {_SAVE_MAX_RETRIES} tentatives — "
        "le fichier est probablement ouvert dans Excel ou verrouillé par la synchronisation cloud."
    ) from last_exc


MAIN_SHEET_NAME = "Appels d'offres"
PPM_SHEET_NAME = "Plans prévisionnels"

MAIN_COLUMNS = [
    "Référence/ID", "Pays", "Source", "Référence", "Objet", "Acheteur",
    "Domaine", "Type de marché", "Procédure", "Lieu d'exécution",
    "Devise", "Montant estimé", "Caution provisoire", "Date publication", "Date limite",
    "Lien avis", "Lien DCE", "Date de collecte",
]
MAIN_HYPERLINK_COLUMNS = {"Lien avis", "Lien DCE"}

PPM_COLUMNS = [
    "Référence/ID", "Pays", "Source", "Ministère", "Autorité contractante",
    "Objet", "Domaine", "Bailleur", "Type de marché", "Mode de passation",
    "Devise", "Montant estimé", "Remarque montant",
    "Date prévisionnelle publication", "Lien source", "Date de collecte",
]
PPM_HYPERLINK_COLUMNS = {"Lien source"}

_HYPERLINK_FONT = Font(color="0563C1", underline="single")


def _ensure_workbook(path: Path) -> Workbook:
    if path.exists():
        return load_workbook(path)
    wb = Workbook()
    wb.remove(wb.active)
    return wb


def _ensure_sheet(wb: Workbook, name: str, columns: list[str]) -> Worksheet:
    if name in wb.sheetnames:
        return wb[name]
    ws = wb.create_sheet(name)
    ws.append(columns)
    return ws


def _index_by_id(ws: Worksheet) -> dict[str, int]:
    index: dict[str, int] = {}
    for row_idx in range(2, ws.max_row + 1):
        value = ws.cell(row=row_idx, column=1).value
        if value:
            index[str(value)] = row_idx
    return index


def _record_to_main_row(record: dict) -> list:
    return [
        record.get("id"),
        record.get("pays"),
        record.get("source"),
        record.get("reference"),
        record.get("objet"),
        record.get("acheteur") or "non précisé",
        record.get("domaine") or "à vérifier",
        record.get("type_marche") or "",
        record.get("procedure_libelle") or "",
        record.get("lieu_execution") or "",
        record.get("devise") or "",
        record.get("montant_estime") or "non précisé",
        record.get("caution_provisoire") or "non précisé",
        record.get("date_publication") or "non précisé",
        record.get("date_limite") or "non précisé",
        record.get("url_avis"),
        record.get("lien_dce"),
        record.get("date_collecte") or dt.date.today().isoformat(),
    ]


def _record_to_ppm_row(record: dict) -> list:
    return [
        record.get("id"),
        record.get("pays"),
        record.get("source"),
        record.get("ministere") or "",
        record.get("acheteur") or "non précisé",
        record.get("objet"),
        record.get("domaine") or "à vérifier",
        record.get("bailleur") or "",
        record.get("type_marche") or "",
        record.get("mode_passation") or "",
        record.get("devise") or "",
        record.get("montant_estime") or "non précisé",
        record.get("montant_remarque", "estimation prévisionnelle — à confirmer"),
        record.get("date_publication") or "non précisé",
        record.get("url_avis"),
        record.get("date_collecte") or dt.date.today().isoformat(),
    ]


def _set_cell(ws: Worksheet, row_idx: int, col_idx: int, value, is_hyperlink_col: bool) -> None:
    cell = ws.cell(row=row_idx, column=col_idx, value=value)
    if is_hyperlink_col and isinstance(value, str) and value.startswith("http"):
        cell.hyperlink = value
        cell.font = _HYPERLINK_FONT


def _upsert_rows(
    ws: Worksheet, records: list[dict], row_builder, columns: list[str], hyperlink_columns: set[str],
    ledger_dir: Path,
) -> tuple[int, int]:
    index = _index_by_id(ws)
    hyperlink_col_idx = {i + 1 for i, name in enumerate(columns) if name in hyperlink_columns}
    num_columns = len(columns)
    # Snapshot AVANT cette exécution — sert uniquement à détecter une
    # suppression manuelle (ID connu mais plus dans l'onglet). `ledger` est
    # la copie de travail mise à jour et sauvegardée en fin de fonction.
    previously_known_ids = _load_ledger(ledger_dir, ws.title)
    ledger = set(previously_known_ids)
    added, updated, ignored_protected, ignored_deleted = 0, 0, 0, 0

    for record in records:
        rec_id = record.get("id")
        if not rec_id:
            continue
        rec_id = str(rec_id)
        ledger.add(rec_id)

        if rec_id in index:
            row_idx = index[rec_id]
            if _is_row_protected(ws, row_idx, num_columns):
                ignored_protected += 1
                continue
            row = row_builder(record)
            for col_idx, value in enumerate(row, start=1):
                _set_cell(ws, row_idx, col_idx, value, col_idx in hyperlink_col_idx)
            updated += 1
        elif rec_id in previously_known_ids:
            # ID géré par une exécution précédente mais absent de l'onglet
            # actuel : supprimé manuellement, jamais réinjecté.
            ignored_deleted += 1
        else:
            row_idx = ws.max_row + 1
            row = row_builder(record)
            for col_idx, value in enumerate(row, start=1):
                _set_cell(ws, row_idx, col_idx, value, col_idx in hyperlink_col_idx)
            index[rec_id] = row_idx
            added += 1

    _save_ledger(ledger_dir, ws.title, ledger)
    if ignored_protected or ignored_deleted:
        logger.info(
            "%s : %d ligne(s) protégée(s) (surlignées, non modifiées), "
            "%d ligne(s) ignorée(s) (supprimées manuellement, non réinjectées)",
            ws.title, ignored_protected, ignored_deleted,
        )
    return added, updated


_STYLE_ATTRS = ("font", "fill", "border", "alignment", "number_format", "protection")


def _copy_cell_style(src_cell, dst_cell) -> None:
    """`copy.copy()` sur chaque objet de style plutôt qu'une affectation
    directe : un style lu sur une cellule (`cell.font`, `cell.fill`...) est
    lié en interne à la table de styles partagée du classeur d'ORIGINE
    (`StyleProxy`) — le réaffecter tel quel, même dans le même classeur,
    lève `TypeError: unhashable type: 'StyleProxy'` (constaté en direct).
    Une copie superficielle détache l'objet de cette liaison avant de
    l'appliquer à la nouvelle cellule."""
    for attr in _STYLE_ATTRS:
        setattr(dst_cell, attr, copy(getattr(src_cell, attr)))
    if src_cell.hyperlink is not None:
        dst_cell.hyperlink = src_cell.hyperlink.target


_SHEET_ORDER = [MAIN_SHEET_NAME, PPM_SHEET_NAME]


def _reorder_sheets(wb: Workbook) -> None:
    wb._sheets.sort(key=lambda ws: _SHEET_ORDER.index(ws.title) if ws.title in _SHEET_ORDER else len(_SHEET_ORDER))


def _rebuild_sheet_dropping_rows(wb: Workbook, sheet_name: str, rows_to_drop: set[int]) -> None:
    """Retire des lignes SANS `ws.delete_rows()` (bug connu : corrompt
    `ws._hyperlinks`, crée des lignes fantômes — cf. README). Reconstruit
    l'onglet dans le MÊME classeur (pas un `Workbook()` séparé) : les objets
    de style (Font/PatternFill...) sont alors directement réutilisables sans
    recréation, contrairement à une copie entre deux classeurs différents."""
    old_ws = wb[sheet_name]
    temp_name = f"__tmp_{sheet_name}"
    new_ws = wb.create_sheet(temp_name)

    new_row_idx = 1
    for row in old_ws.iter_rows():
        if row[0].row in rows_to_drop:
            continue
        for cell in row:
            new_cell = new_ws.cell(row=new_row_idx, column=cell.column, value=cell.value)
            _copy_cell_style(cell, new_cell)
        new_row_idx += 1

    del wb[sheet_name]
    new_ws.title = sheet_name
    _reorder_sheets(wb)


def _remove_stale_rows(
    wb: Workbook, sheet_name: str, columns: list[str], valid_ids: set[str], sources_in_scope: set[str]
) -> int:
    """Retire les lignes d'un onglet qui appartiennent aux `sources_in_scope`
    mais dont l'ID n'est plus dans `valid_ids` (avis clos, reclassé hors IT,
    disparu de la source...). Les lignes PROTÉGÉES (surlignées manuellement)
    sont TOUJOURS conservées, quel que soit leur ID — une décision manuelle
    prime sur la collecte automatique."""
    if sheet_name not in wb.sheetnames or not sources_in_scope:
        return 0
    ws = wb[sheet_name]
    source_col = columns.index("Source") + 1
    num_columns = len(columns)

    to_drop = set()
    for row_idx in range(2, ws.max_row + 1):
        row_id = ws.cell(row=row_idx, column=1).value
        row_source = ws.cell(row=row_idx, column=source_col).value
        if not row_id or row_source not in sources_in_scope:
            continue
        if str(row_id) in valid_ids:
            continue
        if _is_row_protected(ws, row_idx, num_columns):
            continue
        to_drop.add(row_idx)

    if to_drop:
        _rebuild_sheet_dropping_rows(wb, sheet_name, to_drop)
    return len(to_drop)


def write_records(
    records: list[dict], path: Path = config.EXCEL_PATH, source: str | set[str] | None = None
) -> tuple[int, int]:
    """Écrit/met à jour les avis actifs (DGMP-CI, PMMP, ONDA, BCEAO, BAD...)
    dans l'onglet principal. Retourne (nb_ajoutés, nb_mis_à_jour). "Lien
    avis" et "Lien DCE" sont écrits comme hyperliens directement cliquables.

    `source` (optionnel, chaîne ou ensemble de chaînes — ex. une collecte
    combinant "BAD (projets)" et "BAD (corporate)" en un seul appel) déclare
    pour quelle(s) source(s) cet appel fait autorité : toute ligne existante
    de ces sources absente de `records` est retirée (sauf si protégée par
    surlignage). Sans ce paramètre, la portée est déduite de `records`
    lui-même — insuffisant si `records` est vide (ex. une source passe de N
    avis à 0 : impossible de savoir quelle source nettoyer sans le préciser
    explicitement)."""
    wb = _ensure_workbook(path)
    ws = _ensure_sheet(wb, MAIN_SHEET_NAME, MAIN_COLUMNS)
    result = _upsert_rows(ws, records, _record_to_main_row, MAIN_COLUMNS, MAIN_HYPERLINK_COLUMNS, path.parent)

    if source:
        sources_in_scope = {source} if isinstance(source, str) else set(source)
    else:
        sources_in_scope = {r.get("source") for r in records if r.get("source")}
    valid_ids = {str(r["id"]) for r in records if r.get("id")}
    removed = _remove_stale_rows(wb, MAIN_SHEET_NAME, MAIN_COLUMNS, valid_ids, sources_in_scope)
    if removed:
        logger.info("%s : %d ligne(s) obsolète(s) retirée(s) (%s)", MAIN_SHEET_NAME, removed, sources_in_scope)

    _save_with_retry(wb, path)
    return result


def write_ppm_records(
    records: list[dict], path: Path = config.EXCEL_PATH, source: str | set[str] | None = None
) -> tuple[int, int]:
    """Écrit/met à jour les projets PPM prévisionnels dans leur propre
    onglet, séparé des avis actifs. "Lien source" est écrit comme hyperlien
    directement cliquable. `source` : cf. write_records — même principe de
    nettoyage des lignes obsolètes, protection des lignes surlignées comprise."""
    wb = _ensure_workbook(path)
    ws = _ensure_sheet(wb, PPM_SHEET_NAME, PPM_COLUMNS)
    result = _upsert_rows(ws, records, _record_to_ppm_row, PPM_COLUMNS, PPM_HYPERLINK_COLUMNS, path.parent)

    if source:
        sources_in_scope = {source} if isinstance(source, str) else set(source)
    else:
        sources_in_scope = {r.get("source") for r in records if r.get("source")}
    valid_ids = {str(r["id"]) for r in records if r.get("id")}
    removed = _remove_stale_rows(wb, PPM_SHEET_NAME, PPM_COLUMNS, valid_ids, sources_in_scope)
    if removed:
        logger.info("%s : %d ligne(s) obsolète(s) retirée(s) (%s)", PPM_SHEET_NAME, removed, sources_in_scope)

    _save_with_retry(wb, path)
    return result
