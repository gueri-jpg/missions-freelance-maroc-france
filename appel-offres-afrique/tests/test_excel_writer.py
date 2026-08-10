from openpyxl import load_workbook
from openpyxl.styles import PatternFill

import excel_writer


def _record(rid: str, objet: str = "Développement logiciel") -> dict:
    return {
        "id": rid, "pays": "Côte d'Ivoire", "source": "DGMP-CI", "reference": rid,
        "objet": objet, "acheteur": "Ministère X", "domaine": "IT confirmé",
        "type_marche": "PRESTATION", "date_publication": "2026-01-01",
        "date_limite": "2026-12-31", "url_avis": "https://example.test",
        "lien_dce": "https://example.test/dce",
    }


def test_write_records_ajoute_puis_met_a_jour(tmp_path):
    path = tmp_path / "veille.xlsx"

    added, updated = excel_writer.write_records([_record("CI-1")], path=path)
    assert (added, updated) == (1, 0)

    added, updated = excel_writer.write_records([_record("CI-1", objet="Objet modifié")], path=path)
    assert (added, updated) == (0, 1)

    wb = load_workbook(path)
    ws = wb[excel_writer.MAIN_SHEET_NAME]
    assert ws.max_row == 2  # en-tête + 1 ligne (pas de doublon)
    objet_col = excel_writer.MAIN_COLUMNS.index("Objet") + 1
    assert ws.cell(row=2, column=objet_col).value == "Objet modifié"

    lien_avis_col = excel_writer.MAIN_COLUMNS.index("Lien avis") + 1
    lien_dce_col = excel_writer.MAIN_COLUMNS.index("Lien DCE") + 1
    assert ws.cell(row=2, column=lien_avis_col).hyperlink is not None
    assert ws.cell(row=2, column=lien_avis_col).hyperlink.target == "https://example.test"
    assert ws.cell(row=2, column=lien_dce_col).hyperlink is not None
    assert ws.cell(row=2, column=lien_dce_col).hyperlink.target == "https://example.test/dce"


def test_write_ppm_records_onglet_separe(tmp_path):
    path = tmp_path / "veille.xlsx"
    ppm_record = {
        "id": "CI-PPM-1", "pays": "Côte d'Ivoire", "source": "DGMP-CI (PPM prévisionnel)",
        "ministere": "Ministère X", "acheteur": "Direction Y", "objet": "Développement logiciel",
        "domaine": "IT confirmé", "bailleur": "ETAT", "type_marche": "Prestation intellectuelle",
        "mode_passation": "Gré à gré", "date_publication": "17/06/26",
        "url_avis": "https://marchespublics.ci/plan_passation",
    }
    excel_writer.write_ppm_records([ppm_record], path=path)

    wb = load_workbook(path)
    assert excel_writer.PPM_SHEET_NAME in wb.sheetnames
    assert excel_writer.MAIN_SHEET_NAME not in wb.sheetnames

    ws = wb[excel_writer.PPM_SHEET_NAME]
    lien_col = excel_writer.PPM_COLUMNS.index("Lien source") + 1
    assert ws.cell(row=2, column=lien_col).hyperlink is not None
    assert ws.cell(row=2, column=lien_col).hyperlink.target == "https://marchespublics.ci/plan_passation"


def _fill_row(path, sheet_name: str, row_id: str) -> None:
    """Simule un surlignage manuel (n'importe quelle couleur) sur la ligne
    portant cet ID, tel qu'un utilisateur le ferait dans Excel."""
    wb = load_workbook(path)
    ws = wb[sheet_name]
    for row in ws.iter_rows(min_row=2):
        if row[0].value == row_id:
            for cell in row:
                cell.fill = PatternFill(fill_type="solid", fgColor="FFFF00")
            break
    wb.save(path)


def _delete_row(path, sheet_name: str, row_id: str) -> None:
    """Simule une suppression manuelle de ligne dans Excel."""
    wb = load_workbook(path)
    ws = wb[sheet_name]
    for row_idx in range(2, ws.max_row + 1):
        if ws.cell(row=row_idx, column=1).value == row_id:
            ws.delete_rows(row_idx)
            break
    wb.save(path)


def test_write_records_ligne_surlignee_non_modifiee(tmp_path):
    path = tmp_path / "veille.xlsx"
    excel_writer.write_records([_record("CI-1", objet="Objet original")], path=path)
    _fill_row(path, excel_writer.MAIN_SHEET_NAME, "CI-1")

    added, updated = excel_writer.write_records([_record("CI-1", objet="Objet modifié par le pipeline")], path=path)
    assert (added, updated) == (0, 0)

    wb = load_workbook(path)
    ws = wb[excel_writer.MAIN_SHEET_NAME]
    objet_col = excel_writer.MAIN_COLUMNS.index("Objet") + 1
    assert ws.cell(row=2, column=objet_col).value == "Objet original"


def test_write_records_ligne_supprimee_non_reinjectee(tmp_path):
    path = tmp_path / "veille.xlsx"
    excel_writer.write_records([_record("CI-1")], path=path)
    _delete_row(path, excel_writer.MAIN_SHEET_NAME, "CI-1")

    added, updated = excel_writer.write_records([_record("CI-1")], path=path)
    assert (added, updated) == (0, 0)

    wb = load_workbook(path)
    ws = wb[excel_writer.MAIN_SHEET_NAME]
    assert ws.max_row == 1  # seulement l'en-tête, CI-1 non réinjecté


def test_write_records_ligne_obsolete_retiree(tmp_path):
    path = tmp_path / "veille.xlsx"
    excel_writer.write_records([_record("CI-1"), _record("CI-2")], path=path, source="DGMP-CI")

    # CI-2 n'est plus un candidat valide cette fois (ex. reclassé hors IT) :
    # source explicitement en autorité sur DGMP-CI -> doit être retiré.
    added, updated = excel_writer.write_records([_record("CI-1")], path=path, source="DGMP-CI")
    assert (added, updated) == (0, 1)

    wb = load_workbook(path)
    ws = wb[excel_writer.MAIN_SHEET_NAME]
    ids = [ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)]
    assert ids == ["CI-1"]


def test_write_records_ligne_obsolete_protegee_conservee(tmp_path):
    path = tmp_path / "veille.xlsx"
    excel_writer.write_records([_record("CI-1"), _record("CI-2")], path=path, source="DGMP-CI")
    _fill_row(path, excel_writer.MAIN_SHEET_NAME, "CI-2")

    excel_writer.write_records([_record("CI-1")], path=path, source="DGMP-CI")

    wb = load_workbook(path)
    ws = wb[excel_writer.MAIN_SHEET_NAME]
    ids = {ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)}
    assert ids == {"CI-1", "CI-2"}  # CI-2 protégée : jamais retirée automatiquement


def test_write_records_source_explicite_nettoie_meme_si_records_vide(tmp_path):
    path = tmp_path / "veille.xlsx"
    excel_writer.write_records([_record("CI-1")], path=path, source="DGMP-CI")

    # La source retombe à 0 avis (ex. plus aucun candidat IT ce run) :
    # source explicite indispensable ici, records=[] seul ne le permettrait pas.
    excel_writer.write_records([], path=path, source="DGMP-CI")

    wb = load_workbook(path)
    ws = wb[excel_writer.MAIN_SHEET_NAME]
    assert ws.max_row == 1  # CI-1 retiré


def test_write_records_ne_nettoie_pas_les_autres_sources(tmp_path):
    path = tmp_path / "veille.xlsx"
    ci_record = _record("CI-1")
    ma_record = _record("MA-1")
    ma_record["source"] = "PMMP"
    excel_writer.write_records([ci_record, ma_record], path=path)

    # Un run DGMP-CI seul (source explicite) ne doit jamais toucher PMMP.
    excel_writer.write_records([], path=path, source="DGMP-CI")

    wb = load_workbook(path)
    ws = wb[excel_writer.MAIN_SHEET_NAME]
    ids = {ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)}
    assert ids == {"MA-1"}
