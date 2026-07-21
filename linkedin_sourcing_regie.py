# -*- coding: utf-8 -*-
"""
Scraper de missions FREELANCE / RÉGIE dans le secteur BANCAIRE au Maroc (puis France).
Rôles ciblés : PMO, AMOA / AMO / MOA / MOE, Chef de projet, Data / BI / Power BI,
               Data Quality, Business Analyst, Product Owner, Consultant SI...

Source : LinkedIn — endpoint PUBLIC "jobs-guest" (aucune connexion requise).
         => n'utilise PAS votre compte LinkedIn, donc aucun risque pour votre profil.

Sortie : fichier Excel mis en forme comme l'onglet "Maroc" du fichier de sourcing,
         avec la colonne SOURCE en LIEN CLIQUABLE vers l'annonce.
         Les nouvelles missions (jamais vues) sont surlignées en vert.

Usage :
    python linkedin_sourcing_regie.py            # Maroc + France
    python linkedin_sourcing_regie.py maroc      # Maroc uniquement
    python linkedin_sourcing_regie.py france     # France uniquement

Réglages rapides via variables d'environnement (facultatif) :
    LKD_PAGES        nombre de pages par requête   (défaut 2, 10 offres/page)
    LKD_MAX_DETAIL   plafond de fiches détaillées   (défaut 140)
    LKD_QUERY_LIMIT  limite le nb de requêtes        (0 = toutes)
    LKD_REGIE_ONLY   1 = régie/freelance seulement (défaut) ; 0 = tout banque

Dépendances : pip install requests beautifulsoup4 openpyxl lxml
"""

import os
import re
import sys
import json
import time
import random
import datetime as dt
from collections import Counter

import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter

from classifier import (classify_all, strip_accents, normalize_title, has_any,
                        ESN_CDI_ONLY, CLIENT_FINAL_EMPLOYERS)
from sources import collect_all_ats

# ----------------------------------------------------------------------- CONFIG
OUTDIR = os.path.dirname(os.path.abspath(__file__))
SEEN_FILE = os.path.join(OUTDIR, "seen_jobs_linkedin.json")
# Fichier de référence (missions curées manuellement) à fusionner
REFERENCE_FILE = os.path.join(OUTDIR, "Sourcing postes freelance M&F.xlsx")

SEARCH_BASE = "https://www.linkedin.com/jobs-guest/jobs/api/seeMoreJobPostings/search"
JOB_BASE = "https://www.linkedin.com/jobs-guest/jobs/api/jobPosting/"

PAGES_PER_QUERY = int(os.getenv("LKD_PAGES", "2"))
MAX_DETAIL_FETCH = int(os.getenv("LKD_MAX_DETAIL", "300"))
QUERY_LIMIT = int(os.getenv("LKD_QUERY_LIMIT", "0"))
# Nb d'offres servant de "graines" pour lire les OFFRES SIMILAIRES suggerees par
# LinkedIn (cf. similar_cards). 0 = desactive. Chaque graine = 1 requete de plus,
# d'ou une borne : 25 graines ~ 25 requetes pour ~200-900 suggestions brutes.
SIMILAR_SEEDS = int(os.getenv("LKD_SIMILAR_SEEDS", "25"))
# Passe "fraicheur" : re-interroge les memes requetes en ne gardant que les
# offres publiees dans les N derniers jours (parametre LinkedIn f_TPR). 0 = off.
FRESH_DAYS = int(os.getenv("LKD_FRESH_DAYS", "7"))
REGIE_ONLY = os.getenv("LKD_REGIE_ONLY", "1") != "0"
# Retire les offres VIVIER (ouvertes mais anciennes) du fichier. Pour les
# réafficher : LKD_KEEP_VIVIER=1
DROP_VIVIER = os.getenv("LKD_KEEP_VIVIER", "0") == "0"
SLEEP_RANGE = (1.6, 3.4)

USER_AGENTS = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.4 Safari/605.1.15",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:125.0) Gecko/20100101 Firefox/125.0",
]

# Rôles ciblés — au moins un doit apparaître dans le titre OU la description
ROLE_KW = [
    "pmo", "amoa", "amo", "moa", "moe", "chef de projet", "cheffe de projet",
    "data", "business intelligence", " bi ", "power bi", "data analyst",
    "data engineer", "data scientist", "consultant", "business analyst",
    "product owner", "scrum master", "analyste", "maîtrise d'ouvrage",
    "maitrise d'ouvrage", "pilotage", "gouvernance", "data quality",
    "data governance", "data steward", "reporting", "décisionnel", "decisionnel",
]

# Secteur bancaire — au moins un dans titre / entreprise / description
BANK_KW = [
    "banque", "bancaire", "banking", "core banking", "monétique", "monetique",
    "crédit", "credit", "cib", "trade finance", " sab ", "opcvm", "bourse",
    "finance de marché", "finance de marche", "salle des marchés", "swift",
    "moyens de paiement", "paiement", "financement", "risque de crédit",
]

# Signaux régie / freelance FORTS uniquement.
# (On a volontairement retiré "consultant", "mission", "contrat" : ces mots
#  apparaissent dans presque toutes les annonces CDI de banque et créaient
#  des faux positifs. La régie se reconnaît à un signal explicite.)
REGIE_KW = [
    "freelance", "régie", "regie", "en régie", "en regie", "portage",
    "sous-traitance", "sous traitance", "prestation externe",
    "consultant externe", "tjm", "taux journalier",
    "auto-entrepreneur", "auto entrepreneur", "profil freelance",
]
# Seul le type d'emploi "Freelance" vaut signal fort ; "Contrat" est trop
# ambigu (CDD interne) pour être retenu seul.
REGIE_EMPLOYMENT = {"freelance"}

# Employeurs = BANQUES en direct. Une banque qui publie recrute en interne
# (CDI), pas en régie. On exclut donc ces annonces SAUF si le texte contient
# malgré tout un signal régie fort explicite.
BANK_EMPLOYERS = [
    "cih", "société générale", "societe generale", "attijariwafa",
    "banque centrale populaire", "bcp", "crédit du maroc", "credit du maroc",
    "cfg bank", "saham bank", "wafasalaf", "bank of africa", "bmce", "bmci",
    "crédit agricole", "credit agricole", "al barid", "arab bank", "citibank",
    "umnia", "bti bank", "dar al amane",
]

# Cabinets / ESN connus pour placer des consultants EN RÉGIE (secteur banque).
# Une annonce de ces employeurs est classée "régie probable" (niveau 1) même
# sans le mot "freelance". Liste enrichie au fil des découvertes.
KNOWN_REGIE_CABINETS = [
    "brome", "gec", "global experts", "adaptive", "trusted advisors",
    "capfi", "fininfo", "adria", "novancy", "altcode", "consort",
    "africashore", "trusted", "it-adaptive",
]

# TIER1_ONLY = 1 : ne garde que la "régie probable" (haute précision).
# TIER1_ONLY = 0 (défaut) : ajoute les autres cabinets/ESN en "à vérifier".
TIER1_ONLY = os.getenv("LKD_TIER1_ONLY", "0") != "0"

COUNTRIES = {
    "maroc": {"tab": "Maroc", "locations": ["Casablanca, Maroc", "Maroc"]},
    "france": {"tab": "France", "locations": ["Île-de-France, France", "France"]},
}

# Requêtes (rôle + banque). Le suffixe "freelance" biaise le classement LinkedIn
# vers les missions en régie.
QUERIES = [   # base partagée (rôle + banque + génériques régie)
    "PMO banque", "PMO bancaire freelance",
    "AMOA banque", "AMOA bancaire freelance",
    "MOA MOE banque",
    "chef de projet bancaire", "chef de projet banque freelance",
    "consultant data banque", "data analyst banque",
    "business intelligence banque", "power bi banque",
    "data quality banque", "data gouvernance banque",
    "business analyst banque", "product owner banque",
    "consultant freelance banque", "mission freelance bancaire",
    "consultant externe banque", "consultant régie banque",
    # --- PILOTAGE IT : le profil de l'offre ETALON (2026-07-17) -------------
    # L'etalon s'intitule "PMO - pilotage de projets strategiques IT" : le mot
    # "banque" n'est QUE dans le texte. Les requetes "<role> banque" ne
    # pouvaient donc PAS la trouver. On cherche desormais le ROLE seul (le
    # filtre bancaire du classifieur fait le tri ensuite) + les MISSIONS de
    # l'etalon (gouvernance, conduite du changement, COPIL).
    # Mesure du 2026-07-17 : LinkedIn guest est STERILE sur ce profil
    # ("PMO transformation" -> 0 resultat, "chef de projet IT" -> 2 hors sujet)
    # et il coute cher (rate-limit). On ne garde que les variantes qui
    # rapportent ; le gisement PMO regie est sur Free-Work (cf. sources.py),
    # d'ou vient l'offre etalon "PMO - pilotage de projets strategiques IT".
    "chef de projet IT banque", "PMO senior banque",
    "conduite du changement banque", "chef de projet transformation bancaire",
    "gouvernance projet banque", "directeur de projet banque",
]
# Requêtes ciblées par pays : banques qui cherchent un freelance EN DIRECT
BANK_QUERIES = {
    "maroc": ["freelance Attijariwafa", "freelance BCP banque",
              "consultant freelance CIH", "freelance Bank of Africa",
              "freelance banque Casablanca"],
    "france": ["freelance BNP Paribas", "freelance Société Générale banque",
               "consultant freelance Crédit Agricole", "mission freelance Natixis",
               "freelance BPCE banque", "freelance La Banque Postale"],
}

# Colonnes de sortie (identiques à l'onglet "Maroc" du fichier de sourcing).
# La colonne A reste un marqueur (date de mise à jour en A1).
HEADERS = ["", "poste", "mission", "entité", "ville", "exigence", "lieu",
           "contact recrutement", "durée", "SOURCE", "publication"]

WORKMODE = [
    ("hybride", "hybride"), ("télétravail", "télétravail"),
    ("teletravail", "télétravail"), ("remote", "télétravail"),
    ("à distance", "télétravail"), ("a distance", "télétravail"),
    ("full remote", "télétravail"), ("sur site", "sur site"),
    ("sur place", "sur site"), ("présentiel", "sur site"),
    ("presentiel", "sur site"), ("on-site", "sur site"),
]


# ------------------------------------------------------------------------ HTTP
def _headers():
    return {
        "User-Agent": random.choice(USER_AGENTS),
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Language": "fr-FR,fr;q=0.9,en;q=0.8",
    }


def _sleep():
    time.sleep(random.uniform(*SLEEP_RANGE))


def _get(url, params=None, tries=3):
    for attempt in range(tries):
        try:
            r = requests.get(url, params=params, headers=_headers(), timeout=25)
        except requests.RequestException as e:
            print("      ! réseau:", e)
            time.sleep(5 * (attempt + 1))
            continue
        if r.status_code == 200:
            return r.text
        if r.status_code == 429:
            wait = 30 * (attempt + 1)
            print(f"      ! 429 (rate limit LinkedIn) -> pause {wait}s")
            time.sleep(wait)
            continue
        if r.status_code in (400, 404):
            return None
        print("      ! HTTP", r.status_code)
        time.sleep(4 * (attempt + 1))
    return None


# --------------------------------------------------------------------- PARSING
def numeric_id(card):
    if card["id"] and str(card["id"]).isdigit():
        return str(card["id"])
    m = re.search(r"(\d{6,})", card["url"])
    return m.group(1) if m else None


def search_cards(query, location, pages, recent_days=0):
    """recent_days > 0 => filtre LinkedIn `f_TPR` (time posted range), en
    secondes. Verifie le 2026-07-19 : l'endpoint guest l'accepte
    (r86400 = 24 h, r604800 = 7 j).

    C'est LE correctif de fraicheur : la recherche par defaut classe par
    PERTINENCE, pas par date, et on ne lit que 2 pages (20 resultats). Une
    offre publiee ce matin peut donc se classer sous la coupure et etre ratee
    (constat de l'utilisatrice : LinkedIn lui montrait des offres "il y a
    4 heures" absentes de l'Excel). Avec f_TPR, on ne voit QUE le recent.
    """
    cards = []
    for p in range(pages):
        params = {"keywords": query, "location": location, "start": p * 10}
        if recent_days:
            params["f_TPR"] = f"r{int(recent_days) * 86400}"
        txt = _get(SEARCH_BASE, params)
        _sleep()
        if not txt:
            break
        soup = BeautifulSoup(txt, "lxml")
        lis = soup.select("li")
        if not lis:
            break
        got = 0
        for li in lis:
            base = li.select_one("div.base-card") or li
            urn = base.get("data-entity-urn", "") if base else ""
            jid = urn.split(":")[-1] if urn else ""
            a = (li.select_one("a.base-card__full-link")
                 or li.select_one("a[href*='/jobs/view/']"))
            link = a["href"].split("?")[0] if a and a.has_attr("href") else ""
            title = li.select_one("h3.base-search-card__title")
            comp = li.select_one("h4.base-search-card__subtitle")
            loc = li.select_one("span.job-search-card__location")
            tnode = li.select_one("time")
            posted = ""
            if tnode:
                posted = tnode.get("datetime") or tnode.get_text(strip=True)
            if not (title and link):
                continue
            cards.append({
                "id": jid or link,
                "poste": title.get_text(strip=True),
                "entite": comp.get_text(strip=True) if comp else "",
                "ville": loc.get_text(strip=True) if loc else "",
                "url": link,
                "publication": posted,
            })
            got += 1
        if got == 0:
            break
    return cards


def similar_cards(job_url):
    """Offres SIMILAIRES que LinkedIn affiche en bas d'une fiche (idee de
    l'utilisatrice, 2026-07-19 : "elles sont beaucoup plus recentes et
    pourraient convenir").

    ATTENTION : l'endpoint guest `jobPosting/{id}` ne les contient PAS (verifie).
    Il faut la PAGE PUBLIQUE `linkedin.com/jobs/view/...`, qui expose ~37 cartes
    `div.base-card` au meme format que la recherche (id, titre, entite, ville,
    date ISO) -> elles retraversent donc tout le pipeline normal (fiche
    detaillee + classifieur), aucun traitement de faveur.
    """
    if not job_url:
        return []
    txt = _get(job_url.split("?")[0])
    _sleep()
    if not txt:
        return []
    soup = BeautifulSoup(txt, "lxml")
    cards, vus = [], set()
    for base in soup.select("div.base-card"):
        a = base.select_one("a[href*='/jobs/view/']")
        if not a or not a.has_attr("href"):
            continue
        link = a["href"].split("?")[0]
        m = re.search(r"-(\d{8,})", link)
        jid = m.group(1) if m else link
        if jid in vus:
            continue
        vus.add(jid)
        title = base.select_one("h3")
        comp = base.select_one("h4")
        loc = base.select_one("[class*=location]")
        tnode = base.select_one("time")
        if not title:
            continue
        cards.append({
            "id": jid,
            "poste": title.get_text(strip=True),
            "entite": comp.get_text(strip=True) if comp else "",
            "ville": loc.get_text(strip=True) if loc else "",
            "url": link,
            "publication": (tnode.get("datetime") or tnode.get_text(strip=True)) if tnode else "",
        })
    return cards


def fetch_detail(jid):
    txt = _get(JOB_BASE + str(jid))
    _sleep()
    if not txt:
        return {}
    soup = BeautifulSoup(txt, "lxml")
    node = (soup.select_one("div.show-more-less-html__markup")
            or soup.select_one("div.description__text"))
    desc = node.get_text(" ", strip=True) if node else ""
    crit = {}
    for li in soup.select("ul.description__job-criteria-list li"):
        lab = li.select_one("h3")
        val = li.select_one("span")
        if lab and val:
            crit[lab.get_text(strip=True).lower()] = val.get_text(strip=True)

    # --- Champs riches (disponibles sur la page publique "guest") ---
    na = (soup.select_one("figcaption.num-applicants__caption")
          or soup.select_one("span.num-applicants__caption")
          or soup.select_one("[class*='num-applicants__caption']"))
    nb_candidats_txt = na.get_text(strip=True) if na else ""
    pt = soup.select_one("span.posted-time-ago__text")
    posted_relative = pt.get_text(strip=True) if pt else ""
    low = txt.lower()
    # NB : clôture / republication ne sont PAS exposées en accès public
    #      (réservées au mode connecté) — on capture "au mieux".
    cloturee = ("n'accepte plus" in low or "no longer accepting" in low
                or "plus de candidatures" in low)
    republication = ""
    if "republi" in low or "reposted" in low:
        i = low.find("republi") if "republi" in low else low.find("reposted")
        republication = re.sub(r"\s+", " ", txt[i:i + 40]).strip()

    return {"desc": desc, "crit": crit, "nb_candidats_txt": nb_candidats_txt,
            "posted_relative": posted_relative, "cloturee": cloturee,
            "republication": republication}


# ---------------------------------------------------------------- CLASSIFYING
def any_kw(text, kws):
    t = " " + text.lower() + " "
    return any(k in t for k in kws)


def work_mode(text):
    t = text.lower()
    for k, v in WORKMODE:
        if k in t:
            return v
    return "NC"


def duration(text):
    t = text.lower()
    rng = re.search(r"\d{1,2}\s*(?:[-/à]|a)\s*\d{1,2}\s*mois", t)
    if rng:
        return rng.group(0)
    one = re.search(r"\d{1,2}\s*mois", t)
    if one:
        return one.group(0)
    if "longue durée" in t or "longue duree" in t:
        return "longue durée"
    return "NC"


def mission_snippet(desc):
    if not desc:
        return "NC"
    parts = re.split(r"(?<=[\.\!\?])\s", desc.strip())
    snip = (parts[0] if parts else desc).strip()
    return (snip[:160] + "…") if len(snip) > 160 else (snip or "NC")


def is_bank_employer(comp):
    c = (comp or "").lower()
    return any(b in c for b in BANK_EMPLOYERS)


def is_regie_cabinet(comp):
    c = (comp or "").lower()
    return any(b in c for b in KNOWN_REGIE_CABINETS)


def classify(card, detail):
    """Retourne (is_role, is_bank, regie, bank_employer, emp)."""
    title = card["poste"]
    comp = card["entite"]
    desc = detail.get("desc", "")
    emp = ""
    for lab, val in detail.get("crit", {}).items():
        if "type" in lab and "emploi" in lab:
            emp = val
    is_role = any_kw(f"{title} {desc}", ROLE_KW)
    is_bank = any_kw(f"{title} {comp} {desc}", BANK_KW)
    regie = any_kw(f"{title} {desc}", REGIE_KW) or (emp.lower() in REGIE_EMPLOYMENT)
    return is_role, is_bank, regie, is_bank_employer(comp), emp


# -------------------------------------------------------------------- RUN
def _enrich_display(a):
    """Ajoute les colonnes 'photo' (mission/lieu/durée) si absentes."""
    txt = a.get("texte", "")
    a.setdefault("mission", mission_snippet(txt))
    a.setdefault("lieu", work_mode(txt))
    a.setdefault("duree", duration(txt))
    return a


def harvest_country(cc):
    """Collecte multi-sources -> liste d'ANNONCES (LinkedIn + ATS), mise en cache
    (permet de re-classer ensuite hors ligne via `reclass`)."""
    conf = COUNTRIES[cc]
    queries = QUERIES + BANK_QUERIES.get(cc, [])
    if QUERY_LIMIT:
        queries = queries[:QUERY_LIMIT]
    print(f"\n### PAYS : {cc.upper()} — collecte ###")

    # Fiches DEJA lues lors des runs precedents : on les reutilise telles quelles.
    # C'est LE gain de temps : seules les NOUVELLES offres sont telechargees
    # (sinon on refait ~300 requetes/pays a chaque run -> plusieurs heures).
    known = {}
    cache_path = os.path.join(OUTDIR, f"cache_annonces_{cc}.json")
    try:
        with open(cache_path, encoding="utf-8") as f:
            for a in json.load(f):
                u = (a.get("url") or "").split("?")[0].rstrip("/")
                if u and a.get("source") == "LinkedIn":
                    known[u] = a
    except Exception:
        pass

    # 1) LinkedIn (endpoint public)
    raw = {}
    for loc in conf["locations"]:
        for q in queries:
            print(f"  · LinkedIn '{q}' @ {loc}")
            for c in search_cards(q, loc, PAGES_PER_QUERY):
                raw.setdefault(c["id"], c)
    print(f"  {len(raw)} offres LinkedIn (avant fiche détaillée).")

    # 1bis) PASSE FRAICHEUR : les memes requetes, mais restreintes aux offres
    # publiees dans les FRESH_DAYS derniers jours (f_TPR). La recherche normale
    # classe par pertinence sur 2 pages : une offre publiee ce matin peut passer
    # sous la coupure. Ici elle ne peut pas etre ratee -- et on ajoute des
    # requetes de ROLE SEUL (le filtre bancaire trie ensuite), car c'est ainsi
    # que LinkedIn suggerait des PMO frais absents de nos resultats.
    if FRESH_DAYS:
        fresh_q = queries + ["PMO", "chef de projet", "AMOA", "business analyst",
                             "product owner", "data analyst", "chef de projet IT"]
        print(f"  · Passe fraicheur (<= {FRESH_DAYS} j)...")
        neuf = 0
        for loc in conf["locations"]:
            for q in fresh_q:
                for c in search_cards(q, loc, 1, recent_days=FRESH_DAYS):
                    if c["id"] not in raw:
                        raw[c["id"]] = c
                        neuf += 1
        print(f"    +{neuf} offres recentes que la recherche normale ratait.")

    # 1ter) OFFRES SIMILAIRES (cf. similar_cards). On part des offres deja
    # PERTINENTES (role ou banque dans le titre) et on lit les suggestions de
    # LinkedIn, souvent bien plus fraiches que le resultat de recherche.
    # Borne par SIMILAR_SEEDS pour ne pas exploser le nombre de requetes.
    if SIMILAR_SEEDS:
        graines = [c for c in raw.values()
                   if any_kw(f"{c['poste']} {c['entite']}", ROLE_KW)][:SIMILAR_SEEDS]
        print(f"  · Offres similaires LinkedIn (a partir de {len(graines)} offres)...")
        trouve = 0
        for c in graines:
            for s in similar_cards(c.get("url", "")):
                if s["id"] not in raw:
                    raw[s["id"]] = s
                    trouve += 1
        print(f"    +{trouve} offres suggerees par LinkedIn.")

    annonces = []
    fetched = reused = 0
    for card in raw.values():
        pre = f"{card['poste']} {card['entite']}"
        if not (any_kw(pre, ROLE_KW) or any_kw(pre, BANK_KW)):
            continue
        u = (card.get("url") or "").split("?")[0].rstrip("/")
        if u in known:                    # deja lue -> reutilisee, 0 requete
            annonces.append(known[u])
            reused += 1
            continue
        detail = {}
        nid = numeric_id(card)
        if nid and fetched < MAX_DETAIL_FETCH:
            detail = fetch_detail(nid)
            fetched += 1
        annonces.append(to_annonce(card, detail))
    print(f"  LinkedIn : {len(annonces)} annonces "
          f"({fetched} fiches telechargees, {reused} reutilisees du cache).")

    # 2) Sources ATS (cabinets de placement — flux publics)
    print("  · Sources ATS (cabinets)...")
    ats = [_enrich_display(a) for a in collect_all_ats(cc)]
    annonces += ats
    print(f"  TOTAL {len(annonces)} annonces (LinkedIn + {len(ats)} ATS).")

    # FUSION avec le cache precedent : une offre qui disparait des resultats de
    # recherche reste CONNUE. Sinon sa ligne devient "orpheline" dans l'Excel et
    # n'est plus jamais re-filtree (des CDI d'anciens runs y survivaient).
    fusion = {}
    try:
        with open(cache_path, encoding="utf-8") as f:
            for a in json.load(f):
                u = (a.get("url") or "").split("?")[0].rstrip("/")
                if u:
                    fusion[u] = a
    except Exception:
        pass
    for a in annonces:
        u = (a.get("url") or "").split("?")[0].rstrip("/")
        if u:
            fusion[u] = a                 # la version fraiche ecrase l'ancienne
    tout = list(fusion.values())
    with open(cache_path, "w", encoding="utf-8") as f:
        json.dump(tout, f, ensure_ascii=False)
    print(f"  Cache : {len(tout)} annonces connues ({len(annonces)} vues ce run).")
    return tout


def _entite_depuis_url(url):
    """Deduit l'entite du slug LinkedIn '.../view/<titre>-at-<entite>-<jobid>'.
    Filet quand le champ entite est vide au scraping : sinon une offre d'une ESN
    ecartee passe (cas reel 2026-07-20 : "Chargé de mission MOA - Brest" chez
    ASTEK, entite vide -> le filtre ESN ne l'attrapait pas)."""
    m = re.search(r"-at-([a-z0-9-]+?)-\d{6,}", (url or "").lower())
    if not m:
        return ""
    return m.group(1).replace("-", " ").strip()


def to_annonce(card, detail):
    """Transforme (card, detail) scrapés en 'annonce' pour la classification."""
    iso = str(card.get("publication") or "")
    iso = iso[:10] if re.match(r"\d{4}-\d{2}-\d{2}", iso) else ""
    entite = card.get("entite", "") or _entite_depuis_url(card.get("url", ""))
    emp = ""
    for lab, val in (detail.get("crit") or {}).items():
        if "type" in lab and "emploi" in lab:
            emp = val
    desc = detail.get("desc", "")
    return {
        "poste": card.get("poste", ""),
        "entite": entite,
        "ville": card.get("ville", ""),
        "url": (card.get("url", "") or "").split("?")[0],
        "date_pub": iso,
        "posted_relative": detail.get("posted_relative", ""),
        "texte": desc,
        "emploi_label": emp,
        "nb_candidats_txt": detail.get("nb_candidats_txt", ""),
        "cloturee": detail.get("cloturee", False),
        "republication": detail.get("republication", ""),
        "open_confirme": False,          # LinkedIn n'expose pas la clôture
        "source": "LinkedIn",
        # colonnes "photo"
        "mission": mission_snippet(desc),
        "lieu": work_mode(desc),
        "duree": duration(desc),
    }


def process_country(cc, annonces, vues, today):
    """annonces (multi-sources) -> classification complète -> fusion historique."""
    items = classify_all(annonces, today)

    # Horodatage REEL (date + heure) de ce passage.
    now_disp = dt.datetime.now().strftime("%d/%m/%Y %H:%M")
    nouveaux = []
    for a in items:
        url = a["url"]
        rec = vues.get(url)
        if rec:
            a["premiere_detection"] = rec.get("premiere_detection", now_disp)
            a["nouveau"] = False
        else:
            a["premiere_detection"] = now_disp
            a["nouveau"] = True
            if a["verdict"] != "ÉCARTÉE":
                nouveaux.append(a)
        a["derniere_verification"] = now_disp
        vues[url] = {
            "premiere_detection": a["premiere_detection"],
            "derniere_verification": now_disp,
            "verdict": a["verdict"],
            "poste": a["poste"],
            "entite": a["entite"],
        }

    conv = [a for a in items if a["verdict"] != "ÉCARTÉE"]
    c = Counter(a["verdict"] for a in items)
    print(f"\n  [{cc}] {len(items)} annonces classées :")
    for v in ["★★ MATCH CŒUR", "★ À SAISIR", "À CONFIRMER", "VIVIER", "ÉCARTÉE"]:
        if c.get(v):
            print(f"       {c[v]:>3}  {v}")
    print(f"     => {len(conv)} CONVENABLES, dont {len(nouveaux)} nouvelles.")
    for a in nouveaux:
        print(f"  ★ NOUVEAU | {a['verdict']} | {a['entite']} | {a['poste']}")
    return items


# -------------------------------------------------------------------- EXCEL
def fmt_pub(p):
    if not p:
        return "NC"
    m = re.match(r"(\d{4})-(\d{2})-(\d{2})", p)
    if m:
        return f"{m.group(3)}/{m.group(2)}/{m.group(1)}"
    return p


# Colonnes de sortie enrichies (photo + nouvelles colonnes de classification).
HEADERS2 = ["", "poste", "mission", "entité", "ville", "exigence", "lieu",
            "contact recrutement", "durée", "SOURCE", "publication",
            "VERDICT", "type réel", "banque", "cœur métier", "candidats",
            "âge (j)", "fenêtre", "multi-ESN", "question B2B",
            "1re détection", "dern. vérif.", "provenance"]
WIDTHS2 = [12, 30, 40, 22, 15, 13, 12, 16, 11, 14, 12,
           16, 14, 10, 11, 11, 8, 12, 12, 42, 12, 12, 20]

VERDICT_FILL = {
    "★★ MATCH CŒUR": "C6EFCE",   # vert
    "★ À SAISIR": "E2EFDA",       # vert clair
    "★ RÉFÉRENCE": "E4DFEC",      # lavande = issu du fichier de référence
    "À CONFIRMER": "FFF2CC",      # jaune
    "VIVIER": "DDEBF7",           # bleu clair
}
ECARTE_FILL = "E7E6E6"           # gris
NEW_ROW_FILL = "D9EAD3"         # vert très clair pour les nouvelles offres
NEW_ROW_FONT = Font(bold=True, color="1F4E78")


def _pub_display(a):
    iso = a.get("date_pub_iso")
    if iso:
        return fmt_pub(iso)
    return a.get("posted_relative") or "NC"


def _write_sheet(wb, title, rows, today, ecarte=False):
    hdr_fill = PatternFill("solid", fgColor="1F4E78")
    hdr_font = Font(bold=True, color="FFFFFF")
    link_font = Font(color="0563C1", underline="single")
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws = wb.create_sheet(title)
    ws["A1"] = f"MISE A JOUR : {today}"
    ws["A1"].font = Font(bold=True, color="C00000")
    ws["A1"].comment = Comment(
        "VERDICT : ★★ MATCH CŒUR (régie+banque+data, ouverte) · ★ À SAISIR · "
        "À CONFIRMER (question B2B fournie) · VIVIER (besoin récurrent) · "
        "ÉCARTÉE (motif).\nPoste en gras = nouveau depuis le dernier passage.\n"
        "Clôture / republié / 'recrute activement' = non exposés en accès "
        "public LinkedIn ; fraîcheur estimée via âge + nb de candidats.",
        "scraper")
    for ci, h in enumerate(HEADERS2[1:], start=2):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border

    r = 2
    for a in rows:
        is_new = bool(a.get("nouveau"))
        fill_hex = ECARTE_FILL if ecarte else VERDICT_FILL.get(a["verdict"], "FFFFFF")
        if is_new and not ecarte:
            fill_hex = NEW_ROW_FILL
        fill = PatternFill("solid", fgColor=fill_hex)
        verdict_txt = a["verdict"]
        if ecarte and a.get("motif"):
            verdict_txt = f"ÉCARTÉE — {a['motif']}"
        multi = f"OUI ({a['multi_esn_groupe']})" if a.get("multi_esn") else ""
        nb = a.get("nb_candidats_int")
        age = a.get("age_jours")
        source = a.get("source", "LinkedIn")
        src_label = source.replace(" (ATS)", "")
        values = [
            None,                                   # A marqueur
            a["poste"], a.get("mission", "NC"), a["entite"],
            (a.get("ville") or "").upper(), a.get("emploi_label") or "NC",
            a.get("lieu", "NC"), a.get("contact") or "NC", a.get("duree", "NC"),
            src_label, _pub_display(a),
            verdict_txt, a["type"], a["banque"],
            "OUI" if a["coeur_metier"] else "",
            nb if nb is not None else "NC",
            age if age is not None else "NC",
            a["fenetre"], multi, a.get("question_b2b", ""),
            a.get("premiere_detection", today), a.get("derniere_verification", today),
            source,
        ]
        for ci, v in enumerate(values, start=1):
            cell = ws.cell(r, ci, v)
            cell.border = border
            cell.fill = fill
            cell.alignment = Alignment(vertical="top", wrap_text=(ci in (3, 20)))
        if is_new:
            ws.cell(r, 2).font = NEW_ROW_FONT
            ws.cell(r, 2).value = f"★ {ws.cell(r, 2).value}"
            ws.cell(r, 2).fill = fill
            ws.cell(r, 11).font = Font(bold=True, color="1F4E78")
            ws.cell(r, 11).fill = fill
        src = ws.cell(r, 10)
        src.hyperlink = a["url"]
        src.font = link_font
        ws.row_dimensions[r].height = 30
        r += 1

    for i, w in enumerate(WIDTHS2, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "B2"
    ws.auto_filter.ref = f"B1:W{max(1, r - 1)}"
    return ws


# Position des colonnes dans le fichier de référence (les 2 onglets diffèrent).
REF_LAYOUT = {
    "Maroc":  {"start": 2, "poste": 2, "mission": 3, "entite": 4, "ville": 5,
               "exigence": 6, "lieu": 7, "contact": 8, "duree": 9, "source": 10},
    "France": {"start": 3, "poste": 3, "mission": 4, "entite": 5, "ville": 6,
               "exigence": 7, "lieu": 8, "contact": 9, "duree": 10, "source": 12},
}


def load_reference(cc, today_iso):
    """Charge les missions curées du fichier de référence (pré-validées).
    Gère les deux mises en page (Maroc / France). Verdict forcé à '★ RÉFÉRENCE'."""
    tab = COUNTRIES[cc]["tab"]
    lay = REF_LAYOUT.get(tab, REF_LAYOUT["Maroc"])
    try:
        wb = openpyxl.load_workbook(REFERENCE_FILE)
        if tab not in wb.sheetnames:
            return []
        ws = wb[tab]
    except FileNotFoundError:
        print("  ! fichier de référence introuvable — fusion ignorée.")
        return []
    except Exception as e:
        print(f"  ! fichier de référence illisible : {e}")
        return []

    def cell(r, key):
        return ws.cell(r, lay[key]).value

    rows = []
    for r in range(lay["start"], ws.max_row + 1):
        poste = cell(r, "poste")
        ent = cell(r, "entite")
        if not poste and not ent:
            continue
        if str(poste).strip().lower() in ("poste", "none", ""):
            continue
        src = ws.cell(r, lay["source"])
        link = src.hyperlink.target if src.hyperlink else ""
        mission = cell(r, "mission") or ""
        rows.append({
            "poste": str(poste or "").strip(),
            "entite": str(ent or "").strip(),
            "ville": cell(r, "ville") or "",
            "url": link,
            "mission": mission or "NC",
            "exigence": cell(r, "exigence") or "NC",
            "lieu": cell(r, "lieu") or "NC",
            "contact": cell(r, "contact") or "NC",
            "duree": cell(r, "duree") or "NC",
            "date_pub": "",
            "texte": f"{poste or ''} {mission or ''}",
            "emploi_label": "",
            "source": "Réf. M&F",
        })
    if not rows:
        return []
    items = classify_all(rows, today_iso)
    for a in items:                       # pré-validées : on force le verdict
        a["verdict"] = "★ RÉFÉRENCE"
        a["motif"] = ""
        a["nouveau"] = False
        a["derniere_verification"] = today_iso
    print(f"  [{cc}] référence : {len(items)} missions curées chargées.")
    return items


_MERGE_ORDER = {"★★ MATCH CŒUR": 0, "★ À SAISIR": 1, "★ RÉFÉRENCE": 2,
                "À CONFIRMER": 3, "VIVIER": 4}


def _dedup_key(a):
    return (strip_accents(a.get("entite", "")).lower().strip(),
            normalize_title(a.get("poste", "")))


def merge_reference(scraped_conv, ref_items):
    """Fusionne convenables scrapées + missions de référence, SANS doublon.
    Doublon = même cabinet + même intitulé normalisé, OU même URL. La version
    scrapée (enrichie) l'emporte ; les lignes de référence uniques sont ajoutées."""
    seen, urls, merged = set(), set(), []
    for a in scraped_conv:
        k = _dedup_key(a)
        if k in seen:                    # même mission vue sur 2 sources -> 1 seule
            continue
        merged.append(a)
        seen.add(k)
        if a.get("url"):
            urls.add(a["url"].split("?")[0])
    added = 0
    for a in ref_items:
        u = (a.get("url") or "").split("?")[0]
        if _dedup_key(a) in seen or (u and u in urls):
            continue
        merged.append(a)
        seen.add(_dedup_key(a))
        added += 1
    merged.sort(key=lambda a: (_MERGE_ORDER.get(a["verdict"], 9),
                               a["age_jours"] if a.get("age_jours") is not None else 9999))
    print(f"  fusion : {len(scraped_conv)} scrapées + {added} de référence "
          f"(doublons retirés : {len(ref_items) - added}).")
    return merged


def is_convenable(a):
    """Offre à retenir pour les onglets 'sourcing'."""
    v = a["verdict"]
    if v == "ÉCARTÉE":
        return False
    if DROP_VIVIER and v == "VIVIER":
        return False
    return True


def write_excel(path, data_by_country, today):
    """Classeur à 4 onglets :
      - 'Maroc (sourcing)' / 'France (sourcing)' : offres SCRAPÉES convenables
        (verdict != ÉCARTÉE), SANS les missions du fichier de référence.
      - 'Maroc' / 'France' : onglets du fichier de référence, IMPORTÉS TELS QUELS
        (on part du fichier de base pour ne rien modifier de son contenu/mise en forme).
    Les écartées restent tracées dans annonces_vues.json."""
    # 1) On part du fichier de base : ses onglets sont ainsi préservés à l'identique.
    try:
        wb = openpyxl.load_workbook(REFERENCE_FILE)
        base_tabs = list(wb.sheetnames)
    except Exception as e:
        print(f"  ! fichier de référence illisible ({e}) — onglets de base absents.")
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        base_tabs = []

    # 2) On AJOUTE les onglets scrapés (sans toucher aux onglets de base).
    scraped_titles = []
    for cc in COUNTRIES:                        # ordre fixe : Maroc puis France
        if cc not in data_by_country:
            continue
        # Convenables, dédoublonnées (même mission vue sur 2 sources -> 1 seule ;
        # la 1re rencontrée gagne = celle du meilleur verdict, liste déjà triée).
        conv, seen_k = [], set()
        for a in data_by_country[cc]:
            if not is_convenable(a):
                continue
            k = _dedup_key(a)
            if k in seen_k:
                continue
            seen_k.add(k)
            conv.append(a)
        title = f"{COUNTRIES[cc]['tab']} (sourcing)"
        if title in wb.sheetnames:
            wb.remove(wb[title])
        _write_sheet(wb, title, conv, today, ecarte=False)
        scraped_titles.append(title)

    # 3) Réordonner : onglets scrapés d'abord, onglets de base ensuite.
    order = scraped_titles + [t for t in base_tabs if t in wb.sheetnames]
    wb._sheets = [wb[t] for t in order]

    try:
        wb.save(path)
        return path
    except PermissionError:
        alt = path.replace(".xlsx", "_NEW.xlsx")
        wb.save(alt)
        print(f"  ! '{os.path.basename(path)}' est ouvert dans Excel — "
              f"écrit dans '{os.path.basename(alt)}'.")
        return alt


# --------------------------------------------------- MISE A JOUR INCREMENTALE
def _colkey(s):
    """Cle de colonne tolerante aux accents/casse/espaces :
    'cœur métier' -> 'coeurmetier', 'âge (j)' -> 'age'."""
    s = (s or "").lower().replace("œ", "oe").replace("(j)", "")
    return re.sub(r"[^a-z0-9]", "", strip_accents(s))


def _row_values_map(a, today):
    """{cle_colonne_normalisee: valeur} pour une annonce."""
    multi = f"OUI ({a['multi_esn_groupe']})" if a.get("multi_esn") else ""
    nb = a.get("nb_candidats_int")
    age = a.get("age_jours")
    source = a.get("source", "LinkedIn")
    return {
        "poste": a["poste"], "mission": a.get("mission", "NC"),
        "entite": a["entite"], "ville": (a.get("ville") or "").upper(),
        "exigence": a.get("emploi_label") or "NC", "lieu": a.get("lieu", "NC"),
        "contactrecrutement": a.get("contact") or "NC",
        "duree": a.get("duree", "NC"), "source": source.replace(" (ATS)", ""),
        "publication": _pub_display(a), "verdict": a["verdict"],
        "typereel": a["type"], "banque": a["banque"],
        "coeurmetier": "OUI" if a["coeur_metier"] else "",
        "candidats": nb if nb is not None else "NC",
        "age": age if age is not None else "NC",
        "fenetre": a["fenetre"], "multiesn": multi,
        "questionb2b": a.get("question_b2b", ""),
        "1redetection": a.get("premiere_detection", today),
        "dernverif": a.get("derniere_verification", today),
        "provenance": source,
    }


def _header_map(ws):
    """{cle_colonne: index} d'apres la ligne 1 REELLE de l'onglet (donc si tu as
    supprime des colonnes, on respecte ta mise en page)."""
    m = {}
    for c in range(1, (ws.max_column or 1) + 1):
        v = ws.cell(1, c).value
        if v:
            m[_colkey(str(v))] = c
    return m


def _existing_urls(ws, src_col):
    urls = set()
    for r in range(2, (ws.max_row or 1) + 1):
        cell = ws.cell(r, src_col)
        if cell.hyperlink and cell.hyperlink.target:
            urls.add(cell.hyperlink.target.split("?")[0].rstrip("/"))
    return urls


def _delete_rows_safe(ws, a_retirer, src_col):
    """Supprime des lignes SANS casser les liens hypertexte.

    PIEGE openpyxl : delete_rows() decale les VALEURS mais PAS les hyperliens
    (stockes par adresse de cellule) -> les liens se retrouvent colles aux
    mauvaises lignes. On memorise donc les liens, on les purge, on supprime,
    puis on les repose a leur NOUVELLE position.
    """
    if not a_retirer:
        return
    a_retirer = sorted(set(a_retirer))
    liens = {}
    for rr in range(2, (ws.max_row or 1) + 1):
        c = ws.cell(rr, src_col)
        if c.hyperlink:
            liens[rr] = c.hyperlink.target
            c.hyperlink = None                 # purge : sinon il reste en place
    for rr in reversed(a_retirer):
        ws.delete_rows(rr)
    for old_r, target in liens.items():
        if old_r in a_retirer:
            continue
        decalage = sum(1 for s in a_retirer if s < old_r)
        ws.cell(old_r - decalage, src_col).hyperlink = target


def _last_data_row(ws, poste_col):
    last = 1
    for r in range(2, (ws.max_row or 1) + 1):
        if ws.cell(r, poste_col).value not in (None, ""):
            last = r
    return last


def _append_offer(ws, r, a, hmap, today):
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    fill = PatternFill("solid", fgColor=NEW_ROW_FILL)
    link_font = Font(color="0563C1", underline="single")
    vals = _row_values_map(a, today)
    for key, col in hmap.items():
        if key not in vals:
            continue
        v = vals[key]
        if key == "poste":
            v = f"★ {v}"
        cell = ws.cell(r, col, v)
        cell.border = border
        cell.fill = fill
        cell.alignment = Alignment(vertical="top",
                                   wrap_text=(key in ("mission", "questionb2b")))
        if key == "source":
            cell.hyperlink = a["url"]
            cell.font = link_font
        elif key == "poste":
            cell.font = NEW_ROW_FONT
    ws.row_dimensions[r].height = 30


def update_excel(path, data_by_country, today):
    """MET A JOUR le fichier de travail SANS rien ecraser :
      - tes modifications (remplissage, surlignage, notes) sont conservees :
        on ne touche JAMAIS a une ligne deja presente ;
      - une mission que TU as supprimee ne revient pas (deja vue + absente
        du fichier = supprimee volontairement) ;
      - tes colonnes supprimees restent supprimees (on ecrit selon l'en-tete reel) ;
      - seules les NOUVELLES offres sont ajoutees a la suite, marquees ★.
    Si le fichier n'existe pas encore -> creation complete."""
    if not os.path.exists(path):
        return write_excel(path, data_by_country, today)
    try:
        wb = openpyxl.load_workbook(path)
    except Exception as e:
        print(f"  ! fichier illisible ({e}) -> recreation complete.")
        return write_excel(path, data_by_country, today)

    total_add = 0
    for cc in COUNTRIES:
        if cc not in data_by_country:
            continue
        title = f"{COUNTRIES[cc]['tab']} (sourcing)"
        if title not in wb.sheetnames:
            continue
        ws = wb[title]
        hmap = _header_map(ws)
        src_col, poste_col = hmap.get("source"), hmap.get("poste")
        if not src_col or not poste_col:
            print(f"  ! [{title}] en-tete introuvable — onglet ignore.")
            continue
        # 1) NETTOYAGE : retire les lignes devenues NON convenables (CDI seul,
        #    perimee, hors domaine...). On ne juge QUE les offres qu'on connait
        #    encore (presentes dans la classification) — jamais les autres.
        juges = {}
        for a in data_by_country[cc]:
            u = (a.get("url") or "").split("?")[0].rstrip("/")
            if u:
                juges[u] = a
        ent_col = hmap.get("entite")
        a_retirer = []
        for rr in range(2, (ws.max_row or 1) + 1):
            if not ws.cell(rr, poste_col).value:
                continue
            c = ws.cell(rr, src_col)
            a = None
            if c.hyperlink:
                a = juges.get(c.hyperlink.target.split("?")[0].rstrip("/"))
            if a is not None:
                if not is_convenable(a):
                    a_retirer.append(rr)      # jugee par le classifier
            elif ent_col:
                # Ligne INJUGEABLE (orpheline / sans lien) : filet de securite
                # sur le NOM de l'employeur (grosses ESN CDI, clients finaux).
                ent = str(ws.cell(rr, ent_col).value or "")
                if has_any(ent, ESN_CDI_ONLY) or has_any(ent, CLIENT_FINAL_EMPLOYERS):
                    a_retirer.append(rr)
        _delete_rows_safe(ws, a_retirer, src_col)     # preserve les hyperliens
        if a_retirer:
            print(f"  [{title}] {len(a_retirer)} ligne(s) devenue(s) non convenable(s) retiree(s).")

        # 1bis) EFFACEMENT DES ETOILES DE LA MAJ PRECEDENTE (choix utilisatrice
        # 2026-07-20 : "★ = uniquement les offres arrivees a la derniere MAJ").
        # Le ★ etait pose UNE SEULE FOIS a l'insertion et n'etait jamais retire,
        # car on ne reecrit jamais une ligne existante (pour preserver les
        # modifications). Il finissait donc par marquer TOUT le fichier :
        # constate le 20/07 -> 31 lignes etoilees venant de 2 dates differentes.
        # On retire donc le prefixe AVANT d'ajouter les nouvelles lignes.
        # NB : on ne touche QUE le prefixe du poste. Les etoiles de la colonne
        # VERDICT (★★ MATCH CŒUR / ★ A SAISIR) et les couleurs/surlignages
        # restent intacts. La colonne "1re detection" garde l'historique.
        deetoilees = 0
        for rr in range(2, (ws.max_row or 1) + 1):
            v = ws.cell(rr, poste_col).value
            if isinstance(v, str) and v.lstrip().startswith("★"):
                ws.cell(rr, poste_col).value = v.lstrip().lstrip("★").lstrip()
                deetoilees += 1
        if deetoilees:
            print(f"  [{title}] {deetoilees} etoile(s) de la MAJ precedente effacee(s).")

        # 2) Etat du fichier APRES nettoyage
        existing = _existing_urls(ws, src_col)
        r = _last_data_row(ws, poste_col)

        conv, seen_k = [], set()
        for a in data_by_country[cc]:
            if not is_convenable(a):
                continue
            k = _dedup_key(a)
            if k in seen_k:
                continue
            seen_k.add(k)
            conv.append(a)

        # Rafraichit UNIQUEMENT la colonne "dern. verif." des offres encore vues
        # aujourd'hui (une seule cellule : tes autres modifications sont intactes).
        verif_col = hmap.get("dernverif")
        vus = {(a.get("url") or "").split("?")[0].rstrip("/") for a in conv}
        if verif_col:
            for rr in range(2, r + 1):
                c = ws.cell(rr, src_col)
                if c.hyperlink and c.hyperlink.target.split("?")[0].rstrip("/") in vus:
                    ws.cell(rr, verif_col).value = today

        added = 0
        for a in conv:
            u = (a.get("url") or "").split("?")[0].rstrip("/")
            if u in existing:
                continue                     # deja dans le fichier -> on n'y touche pas
            if not a.get("nouveau"):
                continue                     # deja vue avant + absente = TU l'as supprimee
            r += 1
            _append_offer(ws, r, a, hmap, today)
            existing.add(u)
            added += 1
        total_add += added
        ws.cell(1, 1).value = f"MISE A JOUR : {today}"
        print(f"  [{title}] {added} nouvelle(s) offre(s) ajoutee(s) a la suite.")

    try:
        wb.save(path)
        print(f"  => {total_add} nouvelle(s) offre(s) au total (tes modifications conservees).")
        return path
    except PermissionError:
        alt = path.replace(".xlsx", "_NEW.xlsx")
        wb.save(alt)
        print(f"  ! '{os.path.basename(path)}' est ouvert dans Excel — "
              f"ecrit dans '{os.path.basename(alt)}'.")
        return alt


# ------------------------------------------------------------------- HISTORIQUE
# annonces_vues.json : mémorise chaque URL vue (première détection / dernière
# vérification / dernier verdict). On n'écrase jamais premiere_detection.
VUES_FILE = os.path.join(OUTDIR, "annonces_vues.json")


def load_vues():
    try:
        with open(VUES_FILE, encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}


def save_vues(vues):
    with open(VUES_FILE, "w", encoding="utf-8") as f:
        json.dump(vues, f, ensure_ascii=False, indent=1)


# -------------------------------------------------------------------- MAIN
def main():
    args = [a.lower() for a in sys.argv[1:]]
    reclass = False
    if args and args[0] in ("reclass", "recla", "cache", "retune"):
        reclass = True          # re-classe à partir du cache, hors ligne
        args = args[1:]

    arg = args[0] if args else "both"
    if arg in ("both", "tout", "all", "m&f"):
        countries = ["maroc", "france"]
    elif arg in COUNTRIES:
        countries = [arg]
    else:
        print("Argument invalide. Utilisez : [reclass] maroc | france | both")
        return

    today_iso = dt.date.today().isoformat()          # sert au calcul d'anciennete
    # Affichage : date + HEURE REELLE de la mise a jour (cellule A1 de l'Excel).
    today_disp = dt.datetime.now().strftime("%d/%m/%Y %H:%M")
    vues = load_vues()
    out = {}
    for cc in countries:
        if reclass:
            cache = os.path.join(OUTDIR, f"cache_annonces_{cc}.json")
            try:
                with open(cache, encoding="utf-8") as f:
                    annonces = json.load(f)
                print(f"\n### PAYS : {cc.upper()} — reclassement (cache, hors ligne) ###")
            except FileNotFoundError:
                print(f"  ! pas de cache pour {cc}. Lancez d'abord une collecte.")
                annonces = []
        else:
            annonces = harvest_country(cc)
        out[cc] = process_country(cc, annonces, vues, today_iso)

    # Complète le fichier avec les autres pays DÉJÀ en cache (fichier combiné :
    # ex. lancer 'france' garde l'onglet Maroc depuis son cache).
    for cc in COUNTRIES:
        if cc in out:
            continue
        cache = os.path.join(OUTDIR, f"cache_annonces_{cc}.json")
        if os.path.exists(cache):
            with open(cache, encoding="utf-8") as f:
                annonces = json.load(f)
            print(f"\n### PAYS : {cc.upper()} — repris du cache (fichier combiné) ###")
            out[cc] = process_country(cc, annonces, vues, today_iso)

    # UN SEUL fichier de travail, MIS A JOUR sans rien ecraser : tes
    # modifications (remplissage, surlignage, suppressions) sont conservees.
    # LKD_OUT permet d'ecrire dans un fichier SEPARE, pour faire relire un
    # changement de regles sans toucher au fichier de travail de l'utilisatrice
    # (demande du 2026-07-19 : "ne reecris pas l'excel, fais-en un autre, je
    # verifierai les resultats avant fusion"). Dans ce cas l'historique n'est
    # pas commite non plus (cf. le garde-fou juste apres).
    path = os.path.join(OUTDIR, os.getenv("LKD_OUT", "Sourcing_regie_banque.xlsx"))
    written = update_excel(path, out, today_disp)

    # L'historique n'est COMMITE QUE si on a vraiment ecrit dans le fichier de
    # travail. Sinon (Excel ouvert -> repli sur _NEW.xlsx) les offres seraient
    # marquees "deja vues" tout en etant ABSENTES du fichier : au run suivant,
    # la maj incrementale les prendrait pour des lignes supprimees par
    # l'utilisatrice et ne les remettrait JAMAIS. (Bug reel du 2026-07-17.)
    # Idem pour un run de RELECTURE (LKD_OUT) : il ne doit rien consommer, sinon
    # les offres n'arriveraient jamais dans le vrai fichier.
    if os.getenv("LKD_OUT"):
        print("  ! Run de relecture (LKD_OUT) : historique NON enregistre, "
              "le fichier de travail n'est pas touche.")
    elif os.path.abspath(written) == os.path.abspath(path):
        save_vues(vues)
    else:
        print("  ! Historique NON enregistre (fichier verrouille) : ces offres "
              "restent 'nouvelles' et reviendront au prochain run.")

    conv = sum(len([a for a in out[cc] if is_convenable(a)]) for cc in out)
    print(f"\n===> {conv} offres SCRAPÉES convenables (onglets '(sourcing)') "
          f"+ onglets de base intacts, dans :\n     {written}")


if __name__ == "__main__":
    main()
