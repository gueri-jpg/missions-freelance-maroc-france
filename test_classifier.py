# -*- coding: utf-8 -*-
"""
Tests unitaires de la couche classifier — exemples RÉELS tirés du sourcing
(fichier Sourcing_LinkedIn_regie_banque_2026-07-14.xlsx / cache_raw_maroc.json).

Lancer :  python -m unittest test_classifier -v
          (ou simplement : python test_classifier.py)

Correspondance avec la spec :
  L9      -> mission_regie                 : test_regie_broome_data
  L17/L18 -> mission_regie + banque        : test_regie_banque_*
  L32     -> recrutement (carrière)        : test_recrutement_carriere
  L40     -> recrutement (client final)    : test_recrutement_client_final
  L22+L28 -> multi_esn                     : test_multi_esn_data_gouvernance
"""

import unittest
import openpyxl
from classifier import classifier, classify_all, detect_banque
from linkedin_sourcing_regie import _write_sheet, NEW_ROW_FILL

TODAY = "2026-07-15"


class TestType(unittest.TestCase):

    def test_regie_broome_data(self):
        """L9 — BROME 'Expert Data analyst' : 'Pour le compte de notre client
        bancaire' => mission_regie, banque OUI, cœur métier, MATCH CŒUR."""
        a = classifier({
            "poste": "Expert Data analyst",
            "entite": "BROME Consulting & Technology",
            "texte": ("Pour Le Compte De Notre Client Bancaire, Nous Sommes à "
                      "La Recherche D'un Expert Data Analyst Pour Assurer Les "
                      "Missions Suivantes : identifier et extraire des données."),
            "date_pub": "2026-07-10", "emploi_label": "Temps plein",
        }, TODAY)
        self.assertEqual(a["type"], "mission_regie")
        self.assertEqual(a["banque"], "OUI")
        self.assertTrue(a["coeur_metier"])

    def test_regie_banque_amoa(self):
        """L17 — mission_regie + banque (AMOA Corporate Banking, vocab freelance)."""
        a = classifier({
            "poste": "CHEF DE PROJET AMOA CORPORATE BANKING",
            "entite": "BROME Consulting & Technology",
            "texte": ("Mission Freelance pour le compte d'un client bancaire de "
                      "la place. Pilotage AMOA. Démarrage ASAP, longue durée."),
            "date_pub": "2026-06-27", "emploi_label": "Temps plein",
        }, TODAY)
        self.assertEqual(a["type"], "mission_regie")
        self.assertEqual(a["banque"], "OUI")
        self.assertTrue(a["domaine_ok"])

    def test_paiement_monetique_est_un_pole_valide(self):
        """RÈGLE MÉTIER 2026-07-20 (tuteurs) : Paiements & Monétique est un POLE
        VALIDE. Une mission AMOA/PMO/BA/PO/Scrum sur la monétique est DANS le
        perimetre. Seuls les ROLES hors perimetre (dev/QA/prod/support/metier)
        restent exclus, monetique ou pas."""
        base = {"entite": "Cabinet", "emploi_label": "Freelance",
                "date_pub": "2026-07-14",
                "texte": "Mission Freelance pour un client bancaire. TJM."}
        for poste in ["CHEF DE PROJET AMOA MONETIQUE",
                      "Chef de Projet Transverse SEPA - Paiements",
                      "Business Analyst Moyens de Paiement",
                      "PMO Paiement", "Product Owner flux paiement"]:
            a = classifier({**base, "poste": poste}, TODAY)
            self.assertFalse(a["hors_domaine"], f"{poste} = pole valide, a garder")
        # mais les ROLES hors perimetre restent exclus meme en monetique
        for poste, texte in [
                ("Ingénieur de Production Monétique", "run et support production"),
                ("Support Applicatif Monétique", "support applicatif niveau 1"),
                ("Développeur Monétique Java", "développement java")]:
            a = classifier({**base, "poste": poste, "texte": texte}, TODAY)
            self.assertTrue(a["hors_domaine"], f"{poste} = role hors perimetre")

    def test_projet_si_urbanisation_ecarte(self):
        """RÈGLE MÉTIER : « chef de projet » ne suffit pas — SI / urbanisation
        sont hors périmètre."""
        for poste in ["Chef de Projet Urbanisation SI", "Chef de projet SI (H/F)"]:
            a = classifier({
                "poste": poste, "entite": "Cabinet X",
                "texte": "Mission freelance client bancaire. TJM.",
                "date_pub": "2026-07-14", "emploi_label": "Freelance",
            }, TODAY)
            self.assertTrue(a["hors_domaine"], f"{poste} devrait etre hors domaine")

    def test_dev_ecarte_mais_data_gardee(self):
        """RÈGLE MÉTIER : le dev/stack est exclu, MAIS la Data/BI reste le cœur
        métier (un 'Data Engineer Teradata' est gardé, un 'Dev Java' non)."""
        dev = classifier({
            "poste": "Développeur Senior Fullstack Java / React", "entite": "Cabinet X",
            "texte": "Mission freelance client bancaire. TJM.",
            "date_pub": "2026-07-14", "emploi_label": "Freelance"}, TODAY)
        self.assertTrue(dev["hors_domaine"])
        data = classifier({
            "poste": "Data Engineer Teradata (F/H)", "entite": "Cabinet X",
            "texte": "Mission freelance client bancaire. TJM.",
            "date_pub": "2026-07-14", "emploi_label": "Freelance"}, TODAY)
        self.assertFalse(data["hors_domaine"])
        self.assertTrue(data["coeur_metier"])

    def test_offre_de_reference_est_match_coeur(self):
        """L'offre PARFAITE fournie par l'utilisatrice (PMO pilotage IT, banque,
        freelance TJM 466€) doit ressortir en ★★ MATCH CŒUR."""
        items = classify_all([{
            "poste": "PMO – pilotage de projets stratégiques IT",
            "entite": "Mon Consultant Indépendant", "ville": "Paris",
            "url": "https://exemple/ref", "date_pub": "2026-07-10",
            "emploi_label": "Freelance",
            "texte": ("Mission freelance. TJM HT max 466 EUR. 6 mois. Forte "
                      "probabilité de renouvellement. Programme de transformation "
                      "des services IT, notre client du secteur bancaire recherche "
                      "un PMO Senior pour le pilotage d'un projet stratégique. "
                      "Structurer la gouvernance, conduite du changement, animer "
                      "COPIL et COPROJ, reporting, synthèses exécutives."),
        }], TODAY)
        self.assertEqual(items[0]["verdict"], "★★ MATCH CŒUR")

    def test_regie_banque_ref(self):
        """L18 — référence REFxx = signal régie, contexte bancaire (Trusted)."""
        a = classifier({
            "poste": "Consultant PMO Senior REF02",
            "entite": "Trusted Advisors",
            "texte": ("REF02 - Pour une banque internationale à Casablanca, "
                      "nous recherchons un PMO senior. Missions de 3 mois "
                      "renouvelables."),
            "date_pub": "2026-07-01", "emploi_label": "Contrat",
        }, TODAY)
        self.assertEqual(a["type"], "mission_regie")
        self.assertEqual(a["banque"], "OUI")

    def test_recrutement_carriere(self):
        """L32 — texte carrière d'un grand cabinet => recrutement (CDI)."""
        a = classifier({
            "poste": "Consultant Data Senior",
            "entite": "Capgemini",
            "texte": ("Rejoignez nos équipes ! Chez Capgemini, nous cultivons "
                      "notre culture d'excellence. Package attractif, évolution "
                      "de carrière et avantages sociaux."),
            "date_pub": "2026-07-08", "emploi_label": "Temps plein",
        }, TODAY)
        self.assertEqual(a["type"], "recrutement")

    def test_recrutement_client_final(self):
        """L40 — client final (assureur/banque) qui recrute pour lui-même."""
        a = classifier({
            "poste": "Business Analyst",
            "entite": "RMA - Royale Marocaine d'Assurance",
            "texte": ("Nous recherchons un(e) Business Analyst IT passionné(e) "
                      "pour rejoindre nos équipes et piloter la transformation "
                      "de nos systèmes d'information."),
            "date_pub": "2026-07-13", "emploi_label": "Temps plein",
        }, TODAY)
        self.assertEqual(a["type"], "recrutement")

    def test_cdi_seul_ecarte(self):
        """RÈGLE MÉTIER : une annonce en CDI SEUL (salaire/avantages/mutuelle)
        SANS aucune ouverture freelance/B2B est du salariat pur => ÉCARTÉE.
        (Les annonces ouvertes disent presque toujours « CDI / Freelance ».)"""
        items = classify_all([{
            "poste": "Consultant PMO Banque",
            "entite": "XYZ Consulting",
            "texte": ("Mission longue en immersion dans l'équipe de notre banque "
                      "cliente. Poste en CDI, salaire attractif, avantages sociaux "
                      "et mutuelle."),
            "date_pub": "2026-07-10", "emploi_label": "Temps plein",
        }], TODAY)
        self.assertEqual(items[0]["type"], "recrutement")
        self.assertEqual(items[0]["verdict"], "ÉCARTÉE")

    def test_cdi_slash_freelance_garde(self):
        """« CDI / Freelance » = ouverture B2B explicite => gardé en régie."""
        a = classifier({
            "poste": "Consultant PMO Banque",
            "entite": "XYZ Consulting",
            "texte": ("Mission longue chez notre banque cliente. Poste en CDI "
                      "ou Freelance, salaire selon profil, mutuelle."),
            "date_pub": "2026-07-10", "emploi_label": "Temps plein",
        }, TODAY)
        self.assertEqual(a["type"], "mission_regie")

    def test_cdi_ou_freelance_reste_regie(self):
        """« CDI ou freelance » = accessible B2B => mission_regie malgré le CDI."""
        a = classifier({
            "poste": "Chef de projet AMOA banque",
            "entite": "XYZ Consulting",
            "texte": ("Mission longue chez notre client bancaire. Poste ouvert en "
                      "CDI ou freelance selon le profil."),
            "date_pub": "2026-07-10", "emploi_label": "CDI",
        }, TODAY)
        self.assertEqual(a["type"], "mission_regie")

    def test_banque_directe_freelance_gardee(self):
        """Une banque (client final) qui cherche un FREELANCE en direct
        => gardé (mission_regie), pas écarté."""
        a = classifier({
            "poste": "Consultant Data Freelance",
            "entite": "CIH BANK",
            "texte": ("CIH Bank recherche un consultant freelance en régie (TJM) "
                      "pour une mission de 6 mois renouvelables sur la data."),
            "date_pub": "2026-07-12", "emploi_label": "Contrat",
        }, TODAY)
        self.assertEqual(a["type"], "mission_regie")

    def test_banque_directe_cdi_ecartee(self):
        """Une banque sans signal régie = recrutement interne => écarté."""
        items = classify_all([{
            "poste": "Chargé d'études Data", "entite": "BCP Maroc",
            "texte": "Rejoignez nos équipes ! Poste en CDI, avantages sociaux.",
            "date_pub": "2026-07-12", "emploi_label": "Temps plein",
        }], TODAY)
        self.assertEqual(items[0]["verdict"], "ÉCARTÉE")

    def test_label_cdi_nexclut_pas(self):
        """Le label 'Temps plein' n'exclut jamais si le texte dit freelance."""
        a = classifier({
            "poste": "Data Analyst banque",
            "entite": "XYZ Consulting",
            "texte": "Opportunité Freelance pour le compte de notre client bancaire.",
            "date_pub": "2026-07-12", "emploi_label": "Temps plein",
        }, TODAY)
        self.assertEqual(a["type"], "mission_regie")

    def test_banque_direct_ecartee(self):
        """Une banque en direct (Attijariwafa) => recrutement => ÉCARTÉE CDI."""
        items = classify_all([{
            "poste": "Chef de projets Conduite du Changement (H/F)",
            "entite": "Attijariwafa bank",
            "texte": ("Avec son modèle de banque universelle, le groupe "
                      "Attijariwafa bank vous propose de rejoindre nos équipes."),
            "date_pub": "2026-07-13", "emploi_label": "Temps plein",
        }], TODAY)
        self.assertEqual(items[0]["type"], "recrutement")
        self.assertEqual(items[0]["verdict"], "ÉCARTÉE")


class TestDomaineFraicheur(unittest.TestCase):

    def test_hors_domaine_cyber(self):
        a = classifier({
            "poste": "Consultant Cybersécurité",
            "entite": "BROME Consulting & Technology",
            "texte": "Mission freelance cybersécurité pour une banque. Pentest.",
            "date_pub": "2026-07-10", "emploi_label": "Temps plein",
        }, TODAY)
        self.assertTrue(a["hors_domaine"])

    def test_fenetre_age(self):
        a = classifier({"poste": "PMO banque", "entite": "X",
                        "texte": "mission freelance banque", "date_pub": "2026-07-14"}, TODAY)
        self.assertEqual(a["fenetre"], "NOUVEAU")          # 1 j
        b = classifier({"poste": "PMO banque", "entite": "X",
                        "texte": "mission freelance banque", "date_pub": "2026-07-01"}, TODAY)
        self.assertEqual(b["fenetre"], "OUVERTE")          # 14 j
        c = classifier({"poste": "PMO banque", "entite": "X",
                        "texte": "mission freelance banque", "date_pub": "2026-06-01"}, TODAY)
        self.assertEqual(c["fenetre"], "AGEE")             # 44 j, sans republication

    def test_question_b2b_pour_a_confirmer(self):
        a = classifier({
            "poste": "Chef de projet SI", "entite": "Consort Group",
            "texte": "Pilotage de projets pour le secteur bancaire. Core banking.",
            "date_pub": "2026-07-10", "emploi_label": "Temps plein",
        }, TODAY)
        self.assertEqual(a["type"], "a_confirmer")
        self.assertIn("B2B", a["question_b2b"])


class TestExcelHighlighting(unittest.TestCase):

    def test_new_offer_highlighted_in_excel(self):
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        rows = [{
            "poste": "PMO Senior",
            "mission": "Pilotage de la transformation",
            "entite": "Cabinet X",
            "ville": "Casablanca",
            "emploi_label": "Temps plein",
            "lieu": "Hybride",
            "contact": "contact@example.com",
            "duree": "6 mois",
            "source": "LinkedIn",
            "date_pub_iso": "2026-07-10",
            "verdict": "★★ MATCH CŒUR",
            "type": "mission_regie",
            "banque": "OUI",
            "coeur_metier": True,
            "fenetre": "NOUVEAU",
            "multi_esn": False,
            "multi_esn_groupe": "",
            "question_b2b": "",
            "nb_candidats_int": 3,
            "age_jours": 1,
            "url": "https://example.com/offre",
            "nouveau": True,
        }]

        ws = _write_sheet(wb, "Test", rows, "2026-07-16")

        self.assertEqual(ws["B2"].value, "★ PMO Senior")
        self.assertEqual(ws["B2"].fill.fgColor.rgb, "00" + NEW_ROW_FILL)


class TestMultiEsn(unittest.TestCase):

    def test_multi_esn_data_gouvernance(self):
        """L22+L28 — 2 cabinets publient 'CP Data Gouvernance' à <15j.
        (NB : Astek — le cas réel d'origine — est désormais dans ESN_CDI_ONLY
        donc écarté ; on teste ici avec 2 cabinets non exclus.)"""
        items = classify_all([
            {"poste": "Chef de projet Data Gouvernance", "entite": "Cabinet Delta",
             "texte": "Pilotage de la gouvernance des données. Secteur banque.",
             "date_pub": "2026-07-09"},
            {"poste": "Chef de Projet Data Gouvernance (H/F)", "entite": "Consort Group",
             "texte": "Cadrage et pilotage data gouvernance pour un client bancaire.",
             "date_pub": "2026-07-08"},
            {"poste": "Développeur Java", "entite": "Cabinet Omega",   # distracteur
             "texte": "Développement backend banque.", "date_pub": "2026-07-08"},
        ], TODAY)
        g = {i["entite"]: i for i in items}
        self.assertTrue(g["Cabinet Delta"]["multi_esn"])
        self.assertTrue(g["Consort Group"]["multi_esn"])
        self.assertEqual(g["Cabinet Delta"]["multi_esn_groupe"],
                         g["Consort Group"]["multi_esn_groupe"])
        self.assertFalse(g["Cabinet Omega"]["multi_esn"])

    def test_grosse_esn_cdi_ecartee(self):
        """RÈGLE : Onepoint / VISEO / Astek recrutent pour leur effectif
        => écartées, sauf signal freelance/TJM explicite."""
        items = classify_all([{
            "poste": "Architecte Data – Banque", "entite": "Onepoint",
            "texte": "À propos de Onepoint. Nous accompagnons les grandes organisations.",
            "date_pub": "2026-07-15", "emploi_label": "Temps plein",
        }], TODAY)
        self.assertEqual(items[0]["verdict"], "ÉCARTÉE")

    def test_grosse_esn_avec_freelance_explicite_gardee(self):
        """...mais si elle annonce explicitement du freelance/TJM, on garde."""
        a = classifier({
            "poste": "Data Analyst Banque", "entite": "Astek",
            "texte": ("Mission freelance pour le compte de notre client bancaire, "
                      "TJM selon profil."),
            "date_pub": "2026-07-15", "emploi_label": "Contrat",
        }, TODAY)
        self.assertEqual(a["type"], "mission_regie")

    def test_qa_test_recette_ecarte(self):
        """Regle utilisatrice 2026-07-17 : QA / test / recette hors perimetre."""
        for poste in ("Test Lead assurance ISTQB - ALM Octane / HP ALM",
                      "Testeur QA Automation - Robot Framework",
                      "Chargé de recette / homologation SI bancaire"):
            a = classifier({"poste": poste, "entite": "X",
                            "texte": "Mission freelance en régie pour une banque. TJM.",
                            "date_pub": "2026-07-14", "emploi_label": "Freelance"}, TODAY)
            self.assertTrue(a["hors_domaine"], f"devrait etre ecarte : {poste}")

    def test_finance_metier_non_it_ecartee(self):
        """Regle utilisatrice 2026-07-17 : finance metier (non IT) hors perimetre."""
        for poste in ("Liquidity Analyst Consultant", "Consultant Consolidation - Banque",
                      "Contrôleur de gestion banque", "Ingénieur quantitatif risques",
                      "Consultant Formateur IA, Secteur Bancaire H/F",
                      "CHEF DE PROJET IAM - assurances H/F"):
            a = classifier({"poste": poste, "entite": "X",
                            "texte": "Mission freelance en régie pour une banque. TJM.",
                            "date_pub": "2026-07-14", "emploi_label": "Freelance"}, TODAY)
            self.assertTrue(a["hors_domaine"], f"devrait etre ecarte : {poste}")

    def test_entites_ecartees_liste_noire(self):
        """Liste noire d'entites (decision utilisatrice) : TOUTES les missions
        d'une entite disparaissent, y compris les futures. RED TIC ajoute le
        2026-07-20 : le cabinet republie en continu (36 missions dans le cache,
        5 nouvelles rien que dans la MAJ du matin) -> une suppression manuelle
        serait a refaire chaque jour."""
        a = classifier({"poste": "PMO STRATEGIQUE", "entite": "RED TIC",
                        "emploi_label": "Freelance", "date_pub": "2026-07-18",
                        "texte": "Mission freelance en régie pour une banque. TJM."},
                       TODAY)
        self.assertTrue(a["hors_domaine"])
        # une autre entite avec la meme mission reste gardee
        b = classifier({"poste": "PMO STRATEGIQUE", "entite": "Autre Cabinet",
                        "emploi_label": "Freelance", "date_pub": "2026-07-18",
                        "texte": "Mission freelance en régie pour une banque. TJM."},
                       TODAY)
        self.assertFalse(b["hors_domaine"])

    def test_lot_20_juillet_ecarte(self):
        """10 offres rejetees explicitement par l'utilisatrice le 2026-07-20."""
        base = {"emploi_label": "Freelance", "date_pub": "2026-07-14"}
        ecarter = [
            # (poste, entite, texte, ville)
            ("Onboarding Customer", "STHREE", "Mission freelance en régie. "
             "Coordonner les processus d'onboarding, interlocuteur des clients "
             "via Salesforce. Reporting.", "Paris"),
            ("Expert Dataiku senior - grand compte bancaire", "CAT-AMANIA",
             "Mission freelance en régie. Expert Dataiku pour la squad Data.", "Paris"),
            # NB : "reporting réglementaire solvabilité (RWA)" a ete RETIRE de ce
            # lot le 2026-07-20 -> desormais VALIDE (poles Risques/ALM), cf.
            # test_cash_management_et_prudentiel_sont_valides.
            ("Quality manager (H/F)", "ALLEGIS", "Mission freelance en régie. "
             "Manuel qualité et référentiels de gouvernance interne. Banque.", "Paris"),
            ("Administrateur M365 – Support L2 et Run", "GEC", "Mission en régie. "
             "Déploiement Microsoft 365, run et support. Banque.", "Yvelines"),
            ("Senior Regulatory Change Project Manager", "CONSULTING",
             "Mission freelance en régie. Réglementations FATCA/AEOI, DAC 6 et "
             "Section 871(m). Banque.", "Paris"),
            ("Project finance SME — drawdown management (Sao Paulo, Brazil)", "MCI",
             "Mission freelance en régie. Equipe locale à São Paulo, banque.", "Paris"),
            ("INSPECTEUR / INSPECTRICE (F/H) - LYON", "CIC Lyonnaise de Banque",
             "Filiale du Crédit Mutuel. Réseau de 360 agences bancaires.", "Lyon"),
        ]
        for poste, entite, texte, ville in ecarter:
            a = classifier({**base, "poste": poste, "entite": entite,
                            "texte": texte, "ville": ville}, TODAY)
            garde = (not a["hors_domaine"]) and a["type"] != "recrutement" \
                and a["banque"] != "NON"
            self.assertFalse(garde, f"devrait etre ecarte : {poste}")

    def test_dataiku_pas_pris_pour_data(self):
        """'dataiku' contient 'data' : sans exclusion dediee, le stade coeur
        metier le prenait pour du Data/BI."""
        a = classifier({"poste": "Expert Dataiku senior", "entite": "X",
                        "emploi_label": "Freelance", "date_pub": "2026-07-14",
                        "texte": "Mission freelance en régie. Banque."}, TODAY)
        self.assertTrue(a["hors_domaine"])
        # un vrai Data Engineer reste coeur metier
        b = classifier({"poste": "Data Engineer Teradata", "entite": "X",
                        "emploi_label": "Freelance", "date_pub": "2026-07-14",
                        "texte": "Mission freelance en régie. Banque."}, TODAY)
        self.assertFalse(b["hors_domaine"])
        self.assertTrue(b["coeur_metier"])

    def test_geographie_hors_scope(self):
        """France + Maroc seulement. Cas reel : "(Sao Paulo, Brazil)"."""
        from classifier import hors_geographie
        self.assertTrue(hors_geographie("PMO (Sao Paulo, Brazil)", "Paris"))
        self.assertTrue(hors_geographie("Chef de projet", "Londres, UK"))
        self.assertTrue(hors_geographie("PMO banque", "Genève"))
        self.assertFalse(hors_geographie("PMO banque", "Paris, France"))
        self.assertFalse(hors_geographie("PMO banque", "Casablanca, Maroc"))
        # "client international" dans le TEXTE ne doit PAS exclure (on ne teste
        # que titre + ville) : garanti par la signature (pas de texte).

    def test_independant_nu_ne_rouvre_pas_un_cdi(self):
        """'independant' NU attrapait "societe de gestion INDEPENDANTE" et le nom
        "Mon Consultant Independant" -> rouvrait des CDI. Cas Moneta / MPG."""
        moneta = classifier({"poste": "Business Analyst Sénior (Asset Management)",
                             "entite": "Moneta Asset Management", "emploi_label": "Temps plein",
                             "date_pub": "2026-07-14",
                             "texte": "Contrat : CDI Localisation : Paris. Société "
                                      "de gestion de portefeuilles indépendante. Banque."}, TODAY)
        self.assertEqual(moneta["type"], "recrutement")
        # un vrai "consultant independant" reste un signal B2B valable
        vrai = classifier({"poste": "PMO banque", "entite": "Cabinet",
                           "emploi_label": "Temps plein", "date_pub": "2026-07-14",
                           "texte": "Recherche consultant indépendant. Salaire "
                                    "possible ou TJM. Banque, COPIL."}, TODAY)
        self.assertNotEqual(vrai["type"], "recrutement")

    def test_cash_management_et_prudentiel_sont_valides(self):
        """RÈGLE MÉTIER 2026-07-20 : cash management / trésorerie / TMS (pôle ALM)
        ET reporting prudentiel solvabilité/RWA (pôles Risques/ALM) sont VALIDES
        pour un rôle AMOA/PMO/BA/Data. Inverse les exclusions du 19/07 et 20/07."""
        base = {"entite": "X", "emploi_label": "Freelance", "date_pub": "2026-07-14",
                "texte": "Mission freelance en régie pour une banque. TJM."}
        for poste in ("Consultant Senior AMOA Cash Management / Trésorerie – Migration TMS",
                      "CHEF DE PROJET SENIOR CASH MANAGEMENT",
                      "Business Analyst ALM Trésorerie",
                      "Consultant reporting réglementaire solvabilité (RWA)",
                      "AMOA COREP FINREP"):
            a = classifier({**base, "poste": poste}, TODAY)
            self.assertFalse(a["hors_domaine"], f"pole valide, a garder : {poste}")
        # mais les instruments FISCAUX (FATCA/DAC6/871m) restent exclus
        fatca = classifier({**base, "poste": "Senior Regulatory Change PM",
                            "texte": "Mission freelance. Réglementations FATCA/AEOI, "
                                     "DAC 6 et Section 871(m). Banque."}, TODAY)
        self.assertTrue(fatca["hors_domaine"], "FATCA/DAC6/871m doivent rester exclus")

    def test_empilement_certifications_si_ecarte(self):
        """Une annonce qui empile les certifications SI decrit un PROFIL
        certifiant, pas une mission. Cas reel BROME "Expert gouvernance SI et
        PMO" (2026-07-19) : 14 certifications listees, aucune mission decrite."""
        a = classifier({
            "poste": "Expert gouvernance SI et PMO", "entite": "BROME",
            "emploi_label": "Temps plein", "date_pub": "2026-07-12",
            "texte": ("Pour le compte de notre client, nous recherchons un Expert "
                      "gouvernance SI et PMO ayant les compétences suivantes : "
                      "CGEIT, COBIT, TOGAF, BPMN2 Foundation, PMP, PRINCE2, CMMI, "
                      "Lean Six SIGMA, APMG Lean IT, ITIL V4, ITIL OSA, SCRUM, "
                      "ISO/IEC 27001, CISA, COMPTIA Cloud+, CDCP."),
        }, TODAY)
        self.assertTrue(a["hors_domaine"])
        # 1-2 certifications citees dans une vraie mission : on garde
        b = classifier({
            "poste": "PMO Senior banque", "entite": "Cabinet",
            "emploi_label": "Freelance", "date_pub": "2026-07-14",
            "texte": ("Mission freelance en régie. Pilotage du programme, COPIL, "
                      "conduite du changement. La certification PMP est un plus."),
        }, TODAY)
        self.assertFalse(b["hors_domaine"], "offre normale jetee sur 1 certification")

    def test_annonce_agee_nest_pas_rouverte(self):
        """Le proxy "peu de candidats = reactivee" plafonnait a 180 j, alors
        qu'on jette deja le VIVIER des 21 j (preuve pourtant PLUS forte).
        Cas signale : "CHEF DE PROJET AMOA TITRE" (BROME), 142 j -> plafond 45 j."""
        from classifier import fenetre_from_age
        # plafond a 21 j (regle utilisatrice explicite du 2026-07-19)
        self.assertEqual(fenetre_from_age(142, False, False, nb_cand=5), "AGEE")
        self.assertEqual(fenetre_from_age(90, False, False, nb_cand=5), "AGEE")
        self.assertEqual(fenetre_from_age(40, False, False, nb_cand=5), "AGEE")
        self.assertEqual(fenetre_from_age(22, False, False, nb_cand=5), "AGEE")
        # <= 21 j : toujours accepte
        self.assertEqual(fenetre_from_age(21, False, False, nb_cand=5), "OUVERTE")
        self.assertEqual(fenetre_from_age(5, False, False, nb_cand=5), "NOUVEAU")
        # republication EXPLICITE : vraie preuve de fraicheur, non plafonnee
        self.assertEqual(fenetre_from_age(60, False, True, nb_cand=5), "ROUVERTE")

    def test_stage_alternance_cdd_dans_le_texte_ecarte(self):
        """PRIORITE utilisatrice 2026-07-21 : aucune offre stage/alternance/CDD.
        On vise la vraie DECLARATION DE CONTRAT, pas l'usage incident du mot
        ("apprentissage continu", "premiere experience (stage ou alternance)",
        "encadrer un stagiaire"). Piege reel : Banque de France ecrivait
        "Type de recrutement : ‎ Stage" avec un caractere invisible (U+200E)."""
        base = {"entite": "X", "emploi_label": "", "date_pub": "2026-07-18"}
        ecarter = [
            ("Gestionnaire opérations de marché",
             "Type de recrutement : ‎ Stage ‎ Domaine : marché."),
            ("Chef de projet AMOA", "Contrat de stage de 6 mois. Banque."),
            ("AMOA banque", "Poste en alternance sur 12 mois. Banque."),
            ("BA banque", "Convention de stage. Gratification. Banque."),
        ]
        for poste, texte in ecarter:
            a = classifier({**base, "poste": poste, "texte": texte}, TODAY)
            self.assertEqual(a["type"], "recrutement", f"stage/alt non ecarte : {poste}")
        # usages INCIDENTS : on garde
        garder = [
            ("BA banque", "Mission freelance en régie. Apprentissage en continu, "
                          "parcours de formation. Banque."),
            ("BA finance de marché", "Mission freelance banque. Première "
                                     "expérience (ouvert stage et alternance) en BA."),
            ("BA asset management", "Mission freelance. Vous encadrez un BA junior "
                                    "(alternant, stagiaire ou CDI). Banque."),
        ]
        for poste, texte in garder:
            a = classifier({**base, "poste": poste, "texte": texte,
                            "emploi_label": "Contrat"}, TODAY)
            self.assertNotEqual(a["type"], "recrutement", f"usage incident jete : {poste}")

    def test_cdi_explicite_dans_le_texte_ecarte(self):
        """Cas reel OPEN "Chef de Projet IT Bancaire" (2026-07-17) : le texte dit
        "Poste en CDI" + PEE/Tickets Restaurant/RTT/mutuelle, mais etait sauve
        par "prestataire" (qui designait un TIERS : "le prestataire qui developpe
        les programmes"). 'prestataire'/'prestation' ne sont plus des signaux
        B2B, et les avantages CDI sont lus dans le corps du texte."""
        texte = ("Nous accompagnons nos clients grands comptes. Nous travaillons "
                 "chez nos clients. Coordonner avec le prestataire qui developpe "
                 "les programmes de migration. Poste en CDI a pourvoir. "
                 "Avantages : PEE, Tickets Restaurant, RTT, prime d'ete, mutuelle.")
        a = classifier({"poste": "Chef de Projet IT Bancaire H/F", "entite": "OPEN",
                        "emploi_label": "Temps plein", "date_pub": "2026-07-14",
                        "texte": texte}, TODAY)
        self.assertEqual(a["type"], "recrutement")
        # meme sans le mot "CDI", les avantages salaries suffisent
        b = classifier({"poste": "Chef de projet banque", "entite": "ESN",
                        "emploi_label": "Temps plein", "date_pub": "2026-07-14",
                        "texte": "Client bancaire. Avantages : tickets restaurant, "
                                 "RTT, prime de vacances, mutuelle. Le prestataire "
                                 "tiers livre les programmes."}, TODAY)
        self.assertEqual(b["type"], "recrutement")

    def test_autodescription_ne_rouvre_pas_une_annonce_a_salaire(self):
        """"societe de conseil" est une AUTO-DESCRIPTION, pas une offre de B2B :
        elle ne doit pas rouvrir une annonce qui affiche un salaire annuel.
        Mais sans preuve de salariat, elle laisse le doute (=> a_confirmer).
        Cas reels 2026-07-17 : HOUSE OF ABY (35-45 kEUR brut/an) vs Tilencia."""
        salaire = classifier({
            "poste": "Product Owner X/F/H", "entite": "HOUSE OF ABY",
            "emploi_label": "Temps plein", "date_pub": "2026-07-14",
            "texte": ("Société de conseil. Projets liés aux activités bancaires. "
                      "Le salaire proposé pour ce poste se situe entre 35 000 € "
                      "et 45 000 € brut par an."),
        }, TODAY)
        self.assertEqual(salaire["type"], "recrutement")
        sans_salaire = classifier({
            "poste": "PMO banque", "entite": "Tilencia",
            "emploi_label": "Temps plein", "date_pub": "2026-07-14",
            "texte": ("Tilencia est une société de conseil. Notre ADN et nos "
                      "valeurs. Le PMO aura pour mission : structuration du "
                      "pilotage global, COPIL, conduite du changement."),
        }, TODAY)
        self.assertEqual(sans_salaire["type"], "a_confirmer")
        # un vrai B2B (TJM) rouvre bien une annonce qui parle de salaire
        mixte = classifier({
            "poste": "PMO banque", "entite": "Cabinet",
            "emploi_label": "Temps plein", "date_pub": "2026-07-14",
            "texte": "CDI ou freelance. Salaire selon profil, TJM 500 €. Banque.",
        }, TODAY)
        self.assertNotEqual(mixte["type"], "recrutement")

    def test_secteur_sante_ecarte_mais_bancassurance_gardee(self):
        """Un cabinet "probablement banque" (BROME) place aussi dans la SANTE :
        "PMO Senior Projets Sante" (SIH hospitalier) passait en PROBABLE (cas
        signale 2026-07-21). On ecarte le SECTEUR sante, mais PAS la
        bancassurance (pole valide) qui parle d'assurance-vie/prevoyance."""
        sante = detect_banque("PMO Senior Projets Sante",
                              "Projets dans le secteur de la santé, mise en place "
                              "de systèmes d'information hospitaliers (SIH).",
                              "BROME Consulting")
        self.assertEqual(sante, "NON")
        # bancassurance : "assurance" + "distribution" (de produits) ne doit PAS
        # etre pris pour une enumeration de plaquette (seuil releve a 3 secteurs)
        banca = detect_banque("Business Analyst Bancassurance",
                             "Notre client bancassureur : refonte du SI de "
                             "distribution de produits d'assurance-vie et épargne "
                             "au sein du réseau bancaire.", "Cabinet")
        self.assertEqual(banca, "OUI")

    def test_plaquette_esn_nest_pas_une_preuve_bancaire(self):
        """L'enumeration de secteurs d'une plaquette d'ESN ("banques et services
        financiers, energie, transport, luxe, sante") ne prouve PAS que la
        mission est bancaire. Cas reel : "Chef de projet GMAO Senior" (Talan,
        maintenance industrielle) classe banque par cette seule phrase ;
        72 annonces concernees le 2026-07-17."""
        plaquettes = [
            ("Chef de projet GMAO Senior",
             "Talan accompagne ses clients dans des secteurs stratégiques : "
             "banques et services financiers, énergie, transport et mobilité, "
             "luxe et distribution, santé.", "Talan"),
            ("Chef de projet migration",
             "Elle accompagne 130 clients notamment dans les secteurs de la "
             "banque, de l'assurance, de la mutuelle, de l'industrie et du "
             "retail.", "HN Services"),
        ]
        for poste, texte, entite in plaquettes:
            self.assertEqual(detect_banque(poste, texte, entite), "NON",
                             f"plaquette prise pour une preuve : {entite}")
        # Une seule mention HORS enumeration suffit a valider le signal :
        # on raisonne par PHRASE, sinon on perdrait ces vraies missions.
        vraies = [
            ("Chef de projet",
             "Nous intervenons dans les secteurs banque, énergie, transport, "
             "santé. Pour cette mission, notre client bancaire souhaite "
             "renforcer son équipe.", "GrosseESN"),
            ("Data Engineer",
             "Nous couvrons banque, assurance, industrie, santé. La mission "
             "porte sur un datalake pour une grande banque de détail.", "ESN"),
            ("PMO banque", "Secteurs variés : banque, industrie, santé, luxe.", "X"),
        ]
        for poste, texte, entite in vraies:
            self.assertEqual(detect_banque(poste, texte, entite), "OUI",
                             f"vraie mission bancaire perdue : {poste}")

    def test_filet_entites_sans_regie(self):
        """FILET DE SECURITE : l'ensemble des annonces d'une entite EST son ATS.
        Une entite qui publie >=3 annonces sans le moindre signal regie/B2B est
        un recruteur CDI (cas reel : Aubay, dont le site n'a que des rubriques
        CDI et stage/alternance). Mais un cabinet qui decrit de VRAIES missions
        chez un client doit etre epargne (cas reel : CITECH)."""
        from classifier import entites_sans_regie
        recruteur = [{"poste": f"Consultant Data {i}", "entite": "BigESN",
                      "texte": "Rejoignez nos equipes ! Nous accompagnons nos "
                               "clients grands comptes. Immersion totale."}
                     for i in range(4)]
        cabinet = [{"poste": f"Chef de projet {i}", "entite": "VraiCabinet",
                    "texte": "Mission freelance en régie chez notre client "
                             "bancaire. TJM selon profil."} for i in range(4)]
        flags = entites_sans_regie(recruteur + cabinet)
        self.assertIn("bigesn", flags, "recruteur CDI non detecte")
        self.assertNotIn("vraicabinet", flags, "vrai cabinet de regie jete !")
        # sous le seuil : on ne condamne pas sur 2 annonces
        self.assertNotIn("bigesn", entites_sans_regie(recruteur[:2]))

    def test_aupres_de_nos_clients_nest_pas_un_signal_regie(self):
        """'aupres de NOS CLIENTS' (pluriel) = description generique du metier
        d'une ESN dans une annonce CDI. Le SINGULIER reste un vrai signal."""
        esn = classifier({"poste": "Consultant Data Power BI", "entite": "GrosseESN",
                          "emploi_label": "Temps plein", "date_pub": "2026-07-14",
                          "texte": "Aider nos clients à valoriser leurs données "
                                   "auprès de nos clients grands comptes. Banque."},
                         TODAY)
        self.assertNotEqual(esn["type"], "mission_regie")
        vrai = classifier({"poste": "PMO Senior", "entite": "Cabinet",
                           "emploi_label": "Temps plein", "date_pub": "2026-07-14",
                           "texte": "Vous interviendrez auprès de notre client, "
                                    "une banque de premier plan."}, TODAY)
        self.assertEqual(vrai["type"], "mission_regie")

    def test_mots_bancaires_courts_en_mot_entier(self):
        """has_any() compare en SOUS-CHAINE : les mots bancaires courts doivent
        etre matches en MOT ENTIER, sinon 'cib' attrape 'cible', 'bale' attrape
        'globale', 'alm' attrape 'palmares'... (112 faux positifs mesures le
        2026-07-17 : un chef de projets Events chez Accor classe 'banque')."""
        faux = [   # NE DOIVENT PAS etre bancaires
            ("Chef de projets Events", "Nos annonces ne ciblent pas de genre.", "Accor"),
            ("PMO Project Manager Officer", "Adapté aux différentes cibles internes.", "Thales"),
            ("Business Analyst", "Définir le système d'information cible.", "emagine"),
            ("QA Manager", "Holding Enhanced Government Security Accreditation.", "LA Int"),
            ("Chef de projet", "Communication verbale et vision globale.", "MIGSO"),
            ("Data Analyst", "Keyrus classé dans le palmarès Le Point.", "Keyrus"),
        ]
        for poste, texte, entite in faux:
            self.assertEqual(detect_banque(poste, texte, entite), "NON",
                             f"faux positif bancaire : {texte}")
        vrais = [  # DOIVENT rester bancaires
            ("Business Analyst CIB", "Mission au sein de la CIB, front office.", "X"),
            ("Chef de projet", "Risque de crédit et octroi de crédits.", "X"),
            ("Consultant", "Mise en conformité Bâle III et ratios.", "X"),
            ("Consultant ALM", "Gestion ALM et bilan de la banque.", "X"),
            ("CP Leasing", "Solution de crédit-bail pour le client.", "X"),
            ("AMOA", "Déploiement du core banking T24.", "X"),
        ]
        for poste, texte, entite in vrais:
            self.assertEqual(detect_banque(poste, texte, entite), "OUI",
                             f"signal bancaire perdu : {texte}")

    # NB (2026-07-20) : les tests test_paiement_cache_dans_le_texte,
    # test_paiement_produit_ecarte_mais_competence_gardee et
    # test_payment_anglais_ecarte ont ete RETIRES : ils encodaient la regle
    # "paiement exclu", inversee par les tuteurs (Paiements & Monetique = pole
    # valide). Le comportement actuel est verrouille par
    # test_paiement_monetique_est_un_pole_valide.

    def test_metier_non_it_ecarte_mais_credit_immo_garde(self):
        """Fuite reelle du re-scrape 2026-07-17 : 'CHEF DE PROJETS IMMOBILIER'
        (Banque de France) = metier batiment, pas IT -> ecarte.
        PIEGE symetrique : 'Credit Immobilier' EST un domaine IT bancaire."""
        base = {"entite": "X", "emploi_label": "Freelance", "date_pub": "2026-07-14",
                "texte": "Mission freelance en régie pour une banque. TJM."}
        for poste in ("CHEF DE PROJETS IMMOBILIER H/F",
                      "Conducteur de travaux - Banque de France",
                      "Chef de projet communication interne"):
            a = classifier({**base, "poste": poste}, TODAY)
            self.assertTrue(a["hors_domaine"], f"devrait etre ecarte : {poste}")
        for poste in ("Business Analyst Crédit Immobilier",
                      "AMOA Crédit Habitat / Immobilier - Banque",
                      "Chef de projet Prêts Immobiliers"):
            a = classifier({**base, "poste": poste}, TODAY)
            self.assertFalse(a["hors_domaine"], f"ne doit PAS etre ecarte : {poste}")

    def test_transformation_nest_pas_formation(self):
        """PIEGE : ' formation ' (exclu) ne doit PAS attraper 'transformation',
        omnipresent dans les missions bancaires — y compris l'offre de reference."""
        a = classifier({
            "poste": "Chef de projet transformation bancaire",
            "entite": "X", "emploi_label": "Freelance", "date_pub": "2026-07-14",
            "texte": ("Mission freelance en régie. Programme de transformation "
                      "digitale de notre client bancaire. Pilotage, COPIL."),
        }, TODAY)
        self.assertFalse(a["hors_domaine"], "'transformation' confondu avec 'formation'")
        # coeur_metier = Data/BI ; ici c'est du pilotage -> domaine_ok suffit.
        self.assertTrue(a["domaine_ok"])


if __name__ == "__main__":
    unittest.main(verbosity=2)
