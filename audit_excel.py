# -*- coding: utf-8 -*-
"""Audit du fichier de sourcing : relit le TEXTE COMPLET de chaque offre presente
dans l'Excel et ne remonte que les SUSPECTES, avec le motif et l'extrait qui
justifie l'alerte.

Ne modifie rien : c'est une aide a la relecture. Ecrit par-dessus le controle
manuel devenu impraticable sur l'onglet France (100+ lignes).

    python audit_excel.py           # les deux onglets
    python audit_excel.py france    # un seul onglet
"""
import json
import os
import re
import sys

import openpyxl

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import classifier as C                                    # noqa: E402
from classifier import classifier, verdict_of             # noqa: E402

OUTDIR = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.join(OUTDIR, "Sourcing_regie_banque.xlsx")

# --- Controles, calques sur ce que l'utilisatrice verifie a la main ----------
CDI_KW = [
    "poste en cdi", "cdi à pourvoir", "cdi a pourvoir", "recrutement en cdi",
    "tickets restaurant", "ticket restaurant", "titres restaurant",
    "plan d'épargne", " pee ", " rtt ", "prime d'été", "prime de vacances",
    "prime vélo", "comité d'entreprise", "avantages sociaux", "13ème mois",
    "mutuelle", "salaire", "package salarial",
]
FREE_KW = ["freelance", "tjm", "régie", "regie", "portage", "taux journalier",
           "indépendant", "independant"]
AGE_MAX = 45          # au-dela, une annonce n'est plus vraiment "ouverte"


def extrait(texte, motscles, largeur=90):
    """Renvoie l'extrait du texte autour du 1er mot-cle trouve (preuve)."""
    t = " ".join((texte or "").split())
    for k in motscles:
        i = t.lower().find(k.strip().lower())
        if i >= 0:
            return "..." + t[max(0, i - largeur // 2):i + largeur] + "..."
    return ""


def charger_cache():
    """url -> annonce, + titre normalise -> annonce (secours si lien absent)."""
    cache = {}
    for pays in ("france", "maroc"):
        p = os.path.join(OUTDIR, f"cache_annonces_{pays}.json")
        try:
            with open(p, encoding="utf-8") as f:
                for a in json.load(f):
                    u = (a.get("url") or "").split("?")[0].rstrip("/")
                    if u:
                        cache[u] = a
                    cache.setdefault(a.get("poste", "").strip().lower(), a)
        except Exception:
            pass
    return cache


def audit(onglets=None):
    wb = openpyxl.load_workbook(XLSX)
    cache = charger_cache()
    onglets = onglets or ["Maroc (sourcing)", "France (sourcing)"]
    total = 0
    for tab in onglets:
        if tab not in wb.sheetnames:
            continue
        ws = wb[tab]
        last = max((r for r in range(2, ws.max_row + 1) if ws.cell(r, 2).value),
                   default=1)
        alertes = []
        for r in range(2, last + 1):
            titre = str(ws.cell(r, 2).value or "").lstrip("★ ").strip()
            entite = str(ws.cell(r, 4).value or "")
            url = None
            for c in range(1, ws.max_column + 1):
                if ws.cell(r, c).hyperlink:
                    url = ws.cell(r, c).hyperlink.target.split("?")[0].rstrip("/")
                    break
            a = cache.get(url) or cache.get(titre.lower())
            if not a:
                alertes.append((r, titre, entite, "TEXTE INTROUVABLE",
                                "offre absente du cache : impossible a re-verifier"))
                continue
            texte = a.get("texte", "") or ""
            cl = classifier(a)
            verdict, _ = verdict_of(cl)

            # 1. CDI dans le texte sans aucune ouverture freelance
            if C.has_any(texte, CDI_KW) and not C.has_any(f"{titre} {texte}", FREE_KW):
                alertes.append((r, titre, entite, "CDI PROBABLE",
                                extrait(texte, CDI_KW)))
            # 2. annonce trop ancienne
            age = cl.get("age_jours")
            if age is not None and age > AGE_MAX:
                alertes.append((r, titre, entite, f"AGEE ({age} j)",
                                f"publiee le {a.get('date_pub') or '?'}"))
            # 3. le classifieur ne la garderait plus (regles durcies depuis)
            if verdict == "ÉCARTÉE" or cl["hors_domaine"] or cl["banque"] == "NON":
                motif = ("hors perimetre" if cl["hors_domaine"]
                         else "pas bancaire" if cl["banque"] == "NON" else verdict)
                alertes.append((r, titre, entite, "NE PASSE PLUS LES FILTRES", motif))
            # 4. recrutement interne
            if cl["type"] == "recrutement":
                alertes.append((r, titre, entite, "RECRUTEMENT (non regie)", ""))

        print(f"\n{'=' * 78}\n{tab} : {last - 1} offres, {len(alertes)} alerte(s)\n{'=' * 78}")
        if not alertes:
            print("  RAS - rien a signaler.")
        for r, titre, entite, motif, preuve in alertes:
            print(f"  L{r:<4} [{motif}]")
            print(f"        {titre[:62]}  ({entite[:24]})")
            if preuve:
                print(f"        {preuve[:150]}")
        total += len(alertes)
    print(f"\n>>> {total} alerte(s) au total.")
    return total


if __name__ == "__main__":
    args = [a.lower() for a in sys.argv[1:]]
    tabs = None
    if args:
        tabs = [t for t in ["Maroc (sourcing)", "France (sourcing)"]
                if t.split()[0].lower() in args]
    audit(tabs)
