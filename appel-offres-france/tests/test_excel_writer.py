"""Tests de excel_writer.py — dédoublonnage par ID, colonnes DCE vides tant
que la synthèse n'a pas tourné, mise à jour incrémentale."""
from __future__ import annotations

from pathlib import Path

import excel_writer


def _sample_record(**overrides) -> dict:
    record = {
        "id": "24-12345",
        "source": "BOAMP",
        "date_publication": "2024-09-27",
        "date_limite": "2024-10-29",
        "acheteur": "Ville de Test",
        "departement": "75",
        "objet": "Développement d'une application de gestion RH",
        "domaine": "IT confirmé",
        "statut_procedure": "Procédure Adaptée",
        "procedure_libelle": "Procédure Adaptée",
        "accord_cadre": False,
        "nature_categorise_libelle": "Avis de marché/",
        "cpv": ["72212000"],
        "montant_estime": "85000",
        "url_avis": "https://www.boamp.fr/pages/avis/?q=idweb:24-12345",
        "lien_profil_acheteur": "https://www.marches-securises.fr",
        "statut_telechargement": "",
        "statut_synthese": "",
        "remarques": "",
    }
    record.update(overrides)
    return record


def test_write_records_cree_le_fichier_et_les_colonnes(tmp_path: Path):
    path = tmp_path / "veille.xlsx"
    added, updated = excel_writer.write_records([_sample_record()], path=path)

    assert added == 1
    assert updated == 0
    assert path.exists()

    rows = excel_writer.read_main_rows(path=path)
    assert len(rows) == 1
    assert rows[0]["Référence/ID"] == "24-12345"
    assert rows[0]["Domaine"] == "IT confirmé"


def test_colonnes_dce_vides_tant_que_synthese_non_lancee(tmp_path: Path):
    path = tmp_path / "veille.xlsx"
    excel_writer.write_records([_sample_record()], path=path)

    rows = excel_writer.read_main_rows(path=path)
    row = rows[0]
    assert row["Pénalités"] in (None, "")
    assert row["Grille de notation"] in (None, "")
    assert row["Présentiel"] in (None, "")
    assert row["Références exigées"] in (None, "")


def test_dedoublonnage_par_id_met_a_jour_au_lieu_de_dupliquer(tmp_path: Path):
    path = tmp_path / "veille.xlsx"
    excel_writer.write_records([_sample_record()], path=path)

    added, updated = excel_writer.write_records(
        [_sample_record(objet="Objet mis à jour", statut_telechargement="téléchargé")],
        path=path,
    )
    assert added == 0
    assert updated == 1

    rows = excel_writer.read_main_rows(path=path)
    assert len(rows) == 1  # pas de duplication
    assert rows[0]["Objet/Mission"] == "Objet mis à jour"
    assert rows[0]["Statut téléchargement"] == "téléchargé"


def test_update_synthesis_columns_remplit_les_champs_dce(tmp_path: Path):
    path = tmp_path / "veille.xlsx"
    excel_writer.write_records([_sample_record()], path=path)

    synthese = {
        "delai_livraison": "6 mois",
        "calendrier": "Démarrage prévu janvier 2025",
        "penalites": "non précisé",
        "presentiel_exige": "non",
        "references_clients_exigees": "3 références de moins de 3 ans",
        "grille_notation": "Prix 40% / Valeur technique 60%",
    }
    ok = excel_writer.update_synthesis_columns("24-12345", synthese, "ok (gemini)", path=path)
    assert ok is True

    rows = excel_writer.read_main_rows(path=path)
    row = rows[0]
    assert row["Délai/durée"] == "6 mois"
    assert row["Pénalités"] == "non précisé"
    assert row["Statut synthèse"] == "ok (gemini)"


def test_update_synthesis_columns_id_inconnu_renvoie_false(tmp_path: Path):
    path = tmp_path / "veille.xlsx"
    excel_writer.write_records([_sample_record()], path=path)
    ok = excel_writer.update_synthesis_columns("ID-INCONNU", {}, "ok", path=path)
    assert ok is False


def test_montant_absent_marque_non_precise(tmp_path: Path):
    path = tmp_path / "veille.xlsx"
    excel_writer.write_records([_sample_record(montant_estime=None)], path=path)
    rows = excel_writer.read_main_rows(path=path)
    assert rows[0]["Montant estimé"] == "non précisé"


def test_write_approch_records_onglet_separe(tmp_path: Path):
    path = tmp_path / "veille.xlsx"
    excel_writer.write_records([_sample_record()], path=path)

    approch_record = {
        "id": "APPROCH-123",
        "siren_acheteur": 110001013,
        "departement": "75",
        "objet": "Assistance à l'exploitation d'une architecture virtuelle",
        "statut": "Ouvert",
        "procedure_libelle": None,
        "cpv": ["72200000"],
        "montant_estime": "40k - 100k€",
        "montant_remarque": "estimation prévisionnelle — à confirmer",
        "date_previsionnelle_publication": "2026-03-30",
        "date_limite": None,
        "duree_mois": 12,
        "lien_consultation": None,
    }
    added, updated = excel_writer.write_approch_records([approch_record], path=path)
    assert added == 1

    from openpyxl import load_workbook

    wb = load_workbook(path)
    assert excel_writer.MAIN_SHEET_NAME in wb.sheetnames
    assert excel_writer.APPROCH_SHEET_NAME in wb.sheetnames
    ws = wb[excel_writer.APPROCH_SHEET_NAME]
    assert ws.cell(row=2, column=1).value == "APPROCH-123"
    montant_remarque_col = excel_writer.APPROCH_COLUMNS.index("Remarque montant") + 1
    assert ws.cell(row=2, column=montant_remarque_col).value == "estimation prévisionnelle — à confirmer"


def test_update_synthesis_columns_gere_un_champ_imbrique(tmp_path: Path):
    """Le LLM renvoie parfois un objet imbriqué malgré la consigne de champs
    plats (ex. calendrier structuré) — openpyxl ne peut écrire que du texte,
    jamais un dict/list brut : la conversion ne doit pas lever d'exception."""
    path = tmp_path / "veille.xlsx"
    excel_writer.write_records([_sample_record()], path=path)

    synthese = {
        "calendrier": {"limite_remise_offres": "15 mai 2026", "demarrage": "non précisé"},
        "penalites": ["Clause 1", "Clause 2"],
    }
    ok = excel_writer.update_synthesis_columns("24-12345", synthese, "ok (gemini)", path=path)
    assert ok is True

    rows = excel_writer.read_main_rows(path=path)
    assert "15 mai 2026" in rows[0]["Calendrier"]
    assert "Clause 1" in rows[0]["Pénalités"]


def test_update_download_status_bulk_une_seule_sauvegarde(tmp_path: Path):
    """La mise à jour groupée doit modifier plusieurs lignes en une seule
    ouverture/sauvegarde du classeur (moins de conflits sur un dossier
    synchronisé cloud qu'un appel par ligne)."""
    path = tmp_path / "veille.xlsx"
    excel_writer.write_records([
        _sample_record(id="A"),
        _sample_record(id="B"),
        _sample_record(id="C"),
    ], path=path)

    n = excel_writer.update_download_status_bulk(
        {"A": "téléchargé", "B": "connexion requise", "ID-INCONNU": "téléchargé"},
        path=path,
    )
    assert n == 2  # "ID-INCONNU" ignoré silencieusement

    rows = excel_writer.read_main_rows(path=path)
    statuts = {r["Référence/ID"]: r["Statut téléchargement"] for r in rows}
    assert statuts["A"] == "téléchargé"
    assert statuts["B"] == "connexion requise"
    assert statuts["C"] in (None, "")
