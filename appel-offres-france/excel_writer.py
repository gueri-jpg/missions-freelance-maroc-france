"""Écriture/mise à jour incrémentale de l'Excel de veille.

Deux onglets :
- "Appels d'offres" : avis BOAMP (source de référence).
- "Projets à venir (APProch)" : projets d'achats prévisionnels APProch,
  non exhaustifs par nature (saisie volontaire par l'acheteur), montants
  marqués "estimation prévisionnelle — à confirmer".

Dédoublonnage par identifiant (1ère colonne). Une ligne déjà présente est
mise à jour en place plutôt que dupliquée, ce qui permet de relancer la
collecte et la synthèse de façon incrémentale sans perdre les colonnes déjà
renseignées (ex. statut téléchargement, synthèse IA).
"""
from __future__ import annotations

import datetime as dt
import logging
import time
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

import config

logger = logging.getLogger(__name__)

# Le projet peut se trouver dans un dossier synchronisé par un client cloud
# (OneDrive...) — le fichier peut alors être brièvement verrouillé par le
# processus de synchronisation ou par Excel lui-même. On retente avant
# d'abandonner plutôt que de faire échouer tout un traitement par lot pour un
# verrou transitoire.
_SAVE_MAX_RETRIES = 5
_SAVE_RETRY_DELAY_SECONDS = 2.0


def _save_with_retry(wb: Workbook, path: Path) -> None:
    last_exc: OSError | None = None
    for attempt in range(_SAVE_MAX_RETRIES):
        try:
            wb.save(path)
            return
        except (PermissionError, OSError) as exc:
            last_exc = exc
            logger.warning(
                "Fichier Excel verrouillé (tentative %d/%d) : %s — nouvelle tentative dans %.0fs",
                attempt + 1, _SAVE_MAX_RETRIES, exc, _SAVE_RETRY_DELAY_SECONDS,
            )
            time.sleep(_SAVE_RETRY_DELAY_SECONDS)
    raise OSError(
        f"Impossible d'écrire {path} après {_SAVE_MAX_RETRIES} tentatives — "
        "le fichier est probablement ouvert dans Excel ou verrouillé par la synchronisation cloud."
    ) from last_exc

MAIN_SHEET_NAME = "Appels d'offres"
APPROCH_SHEET_NAME = "Projets à venir (APProch)"

MAIN_COLUMNS = [
    "Référence/ID", "Source", "Date publication", "Date limite",
    "Acheteur", "Département", "Objet/Mission", "Domaine",
    "Procédure", "Type de contrat", "CPV", "Montant estimé",
    "Délai/durée", "Calendrier", "Pénalités", "Présentiel",
    "Références exigées", "Grille de notation",
    "Lien avis", "Lien profil acheteur/DCE",
    "Statut téléchargement", "Statut synthèse", "Remarques",
    "Date de collecte",
]

APPROCH_COLUMNS = [
    "Référence/ID", "Acheteur", "SIREN acheteur", "Département", "Objet", "Domaine",
    "Catégorie d'achat", "Statut", "Procédure", "CPV", "Montant estimé", "Remarque montant",
    "Date prévisionnelle publication", "Date cible remise offres",
    "Durée prévisionnelle (mois)", "Lien consultation", "Date de collecte",
]


def _ensure_workbook(path: Path) -> Workbook:
    if path.exists():
        return load_workbook(path)
    wb = Workbook()
    wb.remove(wb.active)
    return wb


def _ensure_sheet(wb: Workbook, name: str, columns: list[str]) -> Worksheet:
    if name in wb.sheetnames:
        return wb[name]
    ws = wb.create_sheet(name)
    ws.append(columns)
    return ws


def _index_by_id(ws: Worksheet) -> dict[str, int]:
    """Retourne {id: numéro de ligne} pour dédoublonnage (colonne 1)."""
    index: dict[str, int] = {}
    for row_idx in range(2, ws.max_row + 1):
        value = ws.cell(row=row_idx, column=1).value
        if value:
            index[str(value)] = row_idx
    return index


def _record_to_main_row(record: dict) -> list:
    synthese = record.get("synthese") or {}
    montant = record.get("montant_estime") or "non précisé"
    type_contrat = "Accord-cadre" if record.get("accord_cadre") else (record.get("nature_categorise_libelle") or "")

    # Pré-remplissage best-effort depuis les données structurées BOAMP
    # (durée, grille de notation) quand elles sont disponibles — fiables
    # car issues directement de l'avis publié, pas d'une invention. La
    # synthèse IA (DCE) prend le relais/affine si elle a tourné.
    delai = synthese.get("delai_livraison") or record.get("duree") or ""
    grille = synthese.get("grille_notation") or record.get("criteres_attribution") or ""
    statut_synthese = record.get("statut_synthese") or ""
    if not statut_synthese and (delai or grille):
        statut_synthese = "pré-rempli (BOAMP) — à confirmer via DCE"

    return [
        record.get("id"),
        record.get("source"),
        record.get("date_publication"),
        record.get("date_limite"),
        record.get("acheteur"),
        record.get("departement"),
        record.get("objet"),
        record.get("domaine") or "à vérifier",
        record.get("statut_procedure") or record.get("procedure_libelle") or "",
        type_contrat,
        ", ".join(record.get("cpv") or []),
        montant,
        delai,
        synthese.get("calendrier", ""),
        synthese.get("penalites", ""),
        synthese.get("presentiel_exige", ""),
        synthese.get("references_clients_exigees", ""),
        grille,
        record.get("url_avis"),
        record.get("lien_profil_acheteur"),
        record.get("statut_telechargement", ""),
        statut_synthese,
        record.get("remarques") or synthese.get("remarques", ""),
        record.get("date_collecte") or dt.date.today().isoformat(),
    ]


def _record_to_approch_row(record: dict) -> list:
    return [
        record.get("id"),
        record.get("acheteur") or "non précisé",
        record.get("siren_acheteur"),
        record.get("departement"),
        record.get("objet"),
        record.get("domaine") or "à vérifier",
        record.get("categorie_achat"),
        record.get("statut"),
        record.get("procedure_libelle"),
        ", ".join(record.get("cpv") or []),
        record.get("montant_estime") or "non précisé",
        record.get("montant_remarque", "estimation prévisionnelle — à confirmer"),
        record.get("date_previsionnelle_publication"),
        record.get("date_limite"),
        record.get("duree_mois"),
        record.get("lien_consultation"),
        record.get("date_collecte") or dt.date.today().isoformat(),
    ]


def _upsert_rows(ws: Worksheet, records: list[dict], row_builder) -> tuple[int, int]:
    index = _index_by_id(ws)
    added, updated = 0, 0
    for record in records:
        rec_id = record.get("id")
        if not rec_id:
            continue
        row = row_builder(record)
        if str(rec_id) in index:
            row_idx = index[str(rec_id)]
            for col_idx, value in enumerate(row, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)
            updated += 1
        else:
            ws.append(row)
            index[str(rec_id)] = ws.max_row
            added += 1
    return added, updated


def write_records(records: list[dict], path: Path = config.EXCEL_PATH) -> tuple[int, int]:
    """Écrit/met à jour les avis BOAMP dans l'onglet principal. Retourne
    (nb_ajoutés, nb_mis_à_jour)."""
    wb = _ensure_workbook(path)
    ws = _ensure_sheet(wb, MAIN_SHEET_NAME, MAIN_COLUMNS)
    result = _upsert_rows(ws, records, _record_to_main_row)
    _save_with_retry(wb, path)
    return result


def write_approch_records(records: list[dict], path: Path = config.EXCEL_PATH) -> tuple[int, int]:
    """Écrit/met à jour les projets prévisionnels APProch dans leur propre
    onglet, séparé des avis officiels BOAMP."""
    wb = _ensure_workbook(path)
    ws = _ensure_sheet(wb, APPROCH_SHEET_NAME, APPROCH_COLUMNS)
    result = _upsert_rows(ws, records, _record_to_approch_row)
    _save_with_retry(wb, path)
    return result


def read_main_rows(path: Path = config.EXCEL_PATH) -> list[dict]:
    """Relit l'onglet principal sous forme de liste de dicts {colonne: valeur},
    utilisé par run_download.py / run_synthesis.py pour repérer les lignes à
    traiter sans dupliquer le modèle de données de collecte."""
    if not path.exists():
        return []
    wb = load_workbook(path)
    if MAIN_SHEET_NAME not in wb.sheetnames:
        return []
    ws = wb[MAIN_SHEET_NAME]
    rows = []
    for row_idx in range(2, ws.max_row + 1):
        values = [ws.cell(row=row_idx, column=c).value for c in range(1, len(MAIN_COLUMNS) + 1)]
        if not values[0]:
            continue
        rows.append(dict(zip(MAIN_COLUMNS, values)))
    return rows


def update_download_status_bulk(status_by_id: dict[str, str], path: Path = config.EXCEL_PATH) -> int:
    """Met à jour la colonne 'Statut téléchargement' pour plusieurs lignes en
    une seule ouverture/sauvegarde du classeur. À privilégier sur un lot
    (ex. run_download.py) plutôt que d'appeler update_download_status en
    boucle : ouvrir/fermer le fichier à chaque ligne multiplie le risque de
    conflit avec la synchronisation cloud (OneDrive) ou une instance Excel
    ouverte. Retourne le nombre de lignes effectivement mises à jour."""
    wb = _ensure_workbook(path)
    if MAIN_SHEET_NAME not in wb.sheetnames:
        return 0
    ws = wb[MAIN_SHEET_NAME]
    index = _index_by_id(ws)
    col_idx = MAIN_COLUMNS.index("Statut téléchargement") + 1

    updated = 0
    for record_id, statut in status_by_id.items():
        row_idx = index.get(str(record_id))
        if row_idx is None:
            continue
        ws.cell(row=row_idx, column=col_idx, value=statut)
        updated += 1

    if updated:
        _save_with_retry(wb, path)
    return updated


def update_download_status(record_id: str, statut: str, path: Path = config.EXCEL_PATH) -> bool:
    """Met à jour uniquement la colonne 'Statut téléchargement' pour une
    ligne existante. Retourne False si l'ID est introuvable. Pour un lot de
    plusieurs lignes, préférer update_download_status_bulk (une seule
    sauvegarde du fichier)."""
    wb = _ensure_workbook(path)
    if MAIN_SHEET_NAME not in wb.sheetnames:
        return False
    ws = wb[MAIN_SHEET_NAME]
    index = _index_by_id(ws)
    row_idx = index.get(str(record_id))
    if row_idx is None:
        return False
    col_idx = MAIN_COLUMNS.index("Statut téléchargement") + 1
    ws.cell(row=row_idx, column=col_idx, value=statut)
    _save_with_retry(wb, path)
    return True


def _to_cell_text(value) -> str:
    """Convertit une valeur de synthèse en texte pour Excel. Le LLM renvoie
    parfois un objet/liste imbriqué malgré la consigne de champs plats
    (ex. calendrier structuré) — openpyxl ne peut écrire que du texte, des
    nombres ou None, jamais un dict/list brut."""
    if value is None:
        return ""
    if isinstance(value, str):
        return value
    if isinstance(value, dict):
        return " ; ".join(f"{k} : {v}" for k, v in value.items() if v not in (None, ""))
    if isinstance(value, list):
        return " ; ".join(_to_cell_text(v) for v in value)
    return str(value)


def update_synthesis_columns(record_id: str, synthese: dict, statut_synthese: str, path: Path = config.EXCEL_PATH) -> bool:
    """Met à jour uniquement les colonnes issues de la synthèse IA pour une
    ligne existante (identifiée par son ID). Retourne False si l'ID est
    introuvable."""
    wb = _ensure_workbook(path)
    if MAIN_SHEET_NAME not in wb.sheetnames:
        return False
    ws = wb[MAIN_SHEET_NAME]
    index = _index_by_id(ws)
    row_idx = index.get(str(record_id))
    if row_idx is None:
        return False

    col_map = {name: i + 1 for i, name in enumerate(MAIN_COLUMNS)}
    ws.cell(row=row_idx, column=col_map["Délai/durée"], value=_to_cell_text(synthese.get("delai_livraison", "")))
    ws.cell(row=row_idx, column=col_map["Calendrier"], value=_to_cell_text(synthese.get("calendrier", "")))
    ws.cell(row=row_idx, column=col_map["Pénalités"], value=_to_cell_text(synthese.get("penalites", "")))
    ws.cell(row=row_idx, column=col_map["Présentiel"], value=_to_cell_text(synthese.get("presentiel_exige", "")))
    ws.cell(row=row_idx, column=col_map["Références exigées"], value=_to_cell_text(synthese.get("references_clients_exigees", "")))
    ws.cell(row=row_idx, column=col_map["Grille de notation"], value=_to_cell_text(synthese.get("grille_notation", "")))
    ws.cell(row=row_idx, column=col_map["Statut synthèse"], value=statut_synthese)
    remarques = synthese.get("remarques")
    if remarques:
        ws.cell(row=row_idx, column=col_map["Remarques"], value=_to_cell_text(remarques))

    _save_with_retry(wb, path)
    return True
