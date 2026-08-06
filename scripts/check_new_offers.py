import openpyxl
from openpyxl.styles.colors import Color
import sys
wb_path = r"c:\Users\PAVILION\Downloads\A explorer\A explorer\Sourcing_regie_banque.xlsx"
try:
    wb = openpyxl.load_workbook(wb_path)
except Exception as e:
    print(f"ERROR loading workbook: {e}")
    sys.exit(2)

sheets = wb.sheetnames
print("Sheets:", sheets)

results = {}

# Helper to check if a cell fill is blue-like
def is_blue_fill(cell):
    try:
        f = cell.fill
        if not f or not hasattr(f, 'fgColor'):
            return False
    except Exception:
        return False
    fg = cell.fill.fgColor
    if fg is None:
        return False
    rgb = None
    if hasattr(fg, 'rgb') and fg.rgb:
        rgb = fg.rgb
    elif hasattr(fg, 'indexed') and fg.indexed:
        # indexed colors are hard to map; consider not blue
        return False
    if not rgb:
        return False
    rgb = rgb.replace('FF','') if rgb.startswith('FF') else rgb
    rgb = rgb.lower()
    # crude check for blue-ish hex
    blue_shades = ['00b0f0','0070c0','00add8','0000ff','c0d9f9','dbe5f1']
    for b in blue_shades:
        if b in rgb:
            return True
    # also check if blue component dominates
    try:
        r = int(rgb[0:2],16)
        g = int(rgb[2:4],16)
        b = int(rgb[4:6],16)
        return b > r and b > g
    except Exception:
        return False

for s in sheets:
    ws = wb[s]
    starred_rows = []
    blue_rows = []
    # iterate rows, skip header row 1
    for r in range(2, ws.max_row+1):
        val = ws.cell(r,2).value
        if val and isinstance(val, str) and val.strip().startswith('★'):
            starred_rows.append(r)
            # check verdict column: try to find header 'VERDICT' to determine column
    # find verdict column index
    verdict_col = None
    for c in range(1, ws.max_column+1):
        h = str(ws.cell(1,c).value or '').strip().upper()
        if 'VERDICT' in h:
            verdict_col = c
            break
    for r in starred_rows:
        blue = False
        if verdict_col:
            cell = ws.cell(r, verdict_col)
            if is_blue_fill(cell):
                blue = True
        else:
            # fallback: check cell 1..ws.max_column for blue
            for c in range(1, ws.max_column+1):
                if is_blue_fill(ws.cell(r,c)):
                    blue = True
                    break
        blue_rows.append((r, blue))
    results[s] = {
        'starred_count': len(starred_rows),
        'starred_rows': [r for r in starred_rows][:20],
        'blue_results_sample': blue_rows[:20]
    }

import json
print(json.dumps(results, indent=2, ensure_ascii=False))
