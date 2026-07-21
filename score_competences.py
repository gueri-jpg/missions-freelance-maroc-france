# -*- coding: utf-8 -*-
"""Classe les offres d'un fichier scrapé par PROXIMITE aux competences idéales
fournies par les tuteurs (2026-07-20) et aux 11 poles metier valides.

Ne modifie rien : sert a "voir ensemble" si le filtre ramene des offres
ressemblant a ce que CFConsulting sait faire.

    python score_competences.py [fichier.xlsx]   (defaut : Sourcing_rescrape_a_voir.xlsx)
"""
import json
import os
import re
import sys
import unicodedata

import openpyxl


def norm(s):
    s = unicodedata.normalize("NFD", str(s or "").lower())
    return "".join(c for c in s if unicodedata.category(c) != "Mn")


# --- Competences idéales (verbatim tuteurs) : chaque bloc = un savoir-faire ----
COMPETENCES = {
    "AMOA / expression de besoin": (3, [
        "amoa", "expression de besoin", "expression des besoins", "recueil des besoins",
        "cadrage", "atelier", "ateliers metier", "parties prenantes"]),
    "Cahier des charges / specs": (3, [
        "cahier des charges", "specification fonctionnelle", "specifications fonctionnelles",
        "spec fonctionnelle", "cartographie fonctionnelle", "user stories", "user story"]),
    "Recette / tests": (3, [
        "recette", "cahier de recette", "test de non regression", "tests de non regression",
        "test metier", "tests metiers", "uat", "qualification"]),
    "Conduite du changement": (3, [
        "conduite du changement", "accompagnement au changement", "formation",
        "refonte de processus", "montee en competence", "adoption"]),
    "Gouvernance / pilotage": (2, [
        "gouvernance", "pilotage", "copil", "coproj", "raci", "comite", "planning",
        "jalons", "risques", "reporting", "instances"]),
    "BI / tableaux de bord": (2, [
        "tableau de bord", "tableaux de bord", "business intelligence", " bi ", "power bi",
        "indicateur", "indicateurs", "kpi", "dataviz", "decisionnel"]),
    "Editeur / integrateur": (1, [
        "editeur", "soutenance", "short list", "shortlist", "integrateur", "appel d'offre",
        "progiciel", "solution du marche"]),
}

# --- 11 poles metier valides --------------------------------------------------
POLES = {
    "Credit & Engagement": ["credit", "engagement", "octroi", "credit-bail", "leasing"],
    "Salle des marches": ["salle des marche", "front office", "middle office", "finance de marche",
                          "trading", "fx", "bonds", "actions", "k+", "kondor", "capital market"],
    "Gestion d'actifs": ["asset management", "gestion d'actif", "gestion d actif", "ocpvm", "opcvm",
                         "portefeuille", "buy side"],
    "Intermediation boursiere": ["intermediation", "bourse", "boursier", "titres", "post-marche",
                                 "post marche", "custody", "depositaire"],
    "Paiements & Monetique": ["paiement", "monetique", "sepa", "carte", "encaissement",
                             "moyens de paiement", "acquiring", "emission", "instant payment"],
    "SIRH": ["sirh", "si rh", "ressources humaines", "paie", "gestion des talents"],
    "Risques": ["risque", "risk", "bale", "credit risk", "risque operationnel", "var "],
    "Planification budgetaire": ["budget", "budgetaire", "planification", "controle de gestion projet",
                                "suivi budgetaire", "pnl", "p&l"],
    "Bancassurance": ["bancassurance", "assurance", "iard", "prevoyance", "epargne"],
    "KYC & Conformite": ["kyc", "conformite", "lcb-ft", "lcbft", "aml", "lutte anti-blanchiment",
                        "due diligence", "reglementaire"],
    "ALM": ["alm", "asset liability", "liquidite", "bilan", "tresorerie", "treasury", "solvabilite",
           "rwa", "corep", "finrep", "cash management"],
}


def score_competences(texte):
    hits = []
    total = 0
    for nom, (poids, kws) in COMPETENCES.items():
        if any(k in texte for k in kws):
            total += poids
            hits.append(nom)
    return total, hits


def poles_de(texte):
    return [nom for nom, kws in POLES.items() if any(k in texte for k in kws)]


def main(path):
    wb = openpyxl.load_workbook(path)
    rows = []
    for tab in ["Maroc (sourcing)", "France (sourcing)"]:
        if tab not in wb.sheetnames:
            continue
        ws = wb[tab]
        last = max((r for r in range(2, ws.max_row + 1) if ws.cell(r, 2).value), default=1)
        for r in range(2, last + 1):
            poste = str(ws.cell(r, 2).value or "").lstrip("★ ").strip()
            entite = str(ws.cell(r, 4).value or "")
            verdict = str(ws.cell(r, 12).value or "")
            neuf = str(ws.cell(r, 2).value or "").startswith("★")
            # texte : titre + mission (col C) pour le score
            blob = norm(poste + " " + str(ws.cell(r, 3).value or ""))
            sc, comps = score_competences(blob)
            pol = poles_de(norm(poste))
            rows.append((sc, tab[:6], neuf, verdict, poste, entite, comps, pol))
    rows.sort(key=lambda x: (-x[0], x[1]))
    print(f"\n{'='*100}\n{path}\n{'='*100}")
    print(f"{'SC':>2} {'N':1} {'pays':6} {'verdict':14} {'poste':46} competences / poles")
    print("-" * 100)
    for sc, pays, neuf, v, poste, entite, comps, pol in rows:
        if sc == 0:
            continue
        tag = "*" if neuf else " "
        cp = ", ".join(c.split(" / ")[0].split(" ")[0] for c in comps[:4])
        print(f"{sc:>2} {tag} {pays:6} {v:14} {poste[:46]:46} {cp} | {'/'.join(pol[:2])}")
    n0 = sum(1 for r in rows if r[0] == 0)
    print(f"\n{len(rows)-n0} offres avec >=1 competence idéale | {n0} sans (score 0)")
    print("(* = nouvelle offre marquee etoile)")


if __name__ == "__main__":
    p = sys.argv[1] if len(sys.argv) > 1 else "Sourcing_rescrape_a_voir.xlsx"
    main(os.path.join(os.path.dirname(os.path.abspath(__file__)), p))
